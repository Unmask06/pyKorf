import logging
import re
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from pykorf import Model
from pykorf.reports.unit_converter import UnitConverter

_logger = logging.getLogger(__name__)


class ResultExporter:
    """A scalable exporter to extract calculated results and input criteria from a pyKorf Model
    and save them into a multi-sheet Excel report or DataFrame dictionary.

    This class correctly relies on the element 'summary' methods to parse the dynamic units
    directly from the underlying KDF records and format them for export.
    """

    def __init__(
        self,
        model: Model,
        basis: str = "",
        remarks: str = "",
        hold: str = "",
        references: list[dict] | None = None,
    ):
        self.model = model
        self._basis = basis
        self._remarks = remarks
        self._hold = hold
        self._references = references or []

        self._converter = UnitConverter()

        # ---------------------------------------------------------
        # REGISTRY: Add new element extractors here in the future
        # ---------------------------------------------------------
        self._extractors = {
            "Feeds": self._extract_feeds,
            "Products": self._extract_products,
            "Pipes": self._extract_pipes,
            "Pumps": self._extract_pumps,
            "Compressors": self._extract_compressors,
            "Valves": self._extract_valves,
            "Heat Exchangers": self._extract_heat_exchangers,
            "Junctions": self._extract_junctions,
            "Misc Equipment": self._extract_misc,
        }

        # Header parsing pattern
        self._header_pattern = re.compile(r"^(.*?) \[(.*?)\]$")

        # Reuseable Styles
        self._styles = {
            "model_title": Font(bold=True, size=18, color="003366"),
            "title": Font(bold=True, size=14, color="003366"),
            "header": Font(bold=True, size=10),
            "unit": Font(italic=True, color="555555", size=10),
            "data": Font(size=10),
            "footer": Font(italic=True, size=9, color="888888"),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "thin_side": Side(style="thin"),
            "thick_side": Side(style="medium"),
        }

    def generate_dataframes(self) -> dict[str, pd.DataFrame]:
        """Runs all registered extractors and returns a dictionary of DataFrames."""
        dfs = {}
        for sheet_name, extractor_func in self._extractors.items():
            data = extractor_func()
            if data:  # Only add the sheet if there are actual elements
                converted_data = self._converter.convert_summary(data)
                dfs[sheet_name] = pd.DataFrame(converted_data)
        return dfs

    def export_to_excel(
        self,
        output_path: str,
        sheet_name: str = "Model Summary",
        elements: list[str] | None = None,
        template_path: str | None = None,
    ) -> str:
        """Generates the DataFrames and writes them sequentially to a single formatted Excel sheet."""
        from pathlib import Path

        from openpyxl import load_workbook

        source_name = Path(self.model._parser.path).name
        _logger.info("── Generate Report ── %s", source_name)
        dfs_flat = self.generate_dataframes()
        non_empty = [k for k, v in dfs_flat.items() if not v.empty]
        _logger.info("   Sections: %s", ", ".join(non_empty) if non_empty else "none")

        if template_path and Path(template_path).exists():
            workbook = load_workbook(template_path)
            worksheet = workbook.active
            worksheet.title = sheet_name
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name

            # Set page setup for A3 Landscape
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.scale = 80

        # Insert References & Design Basis sheet as the first sheet (if data present)
        if self._basis or self._remarks or self._hold or self._references:
            self._write_references_sheet(workbook)

        # Row 1 — Model title from KORF SYMBOL (FSIZ=2)
        model_title = self._get_model_title()
        if model_title:
            title_cell = worksheet.cell(row=1, column=1, value=model_title)
            title_cell.font = self._styles["model_title"]
            worksheet.row_dimensions[1].height = 28

        # Row 2 — Source file
        source_name = Path(self.model._parser.path).name
        worksheet.cell(row=2, column=1, value=f"Source File: {source_name}").font = self._styles[
            "header"
        ]

        # Row 3 — Cases
        case_names = self.model.general.case_descriptions if hasattr(self.model, "general") else []
        if case_names:
            worksheet.cell(
                row=3, column=1, value=f"Cases: {'; '.join(case_names)}"
            ).font = self._styles["header"]

        current_row_left = 8 if template_path else 5
        current_row_right = 8 if template_path else 5

        element_keys = elements if elements is not None else list(self._extractors.keys())

        for element_type in element_keys:
            if element_type not in dfs_flat or dfs_flat[element_type].empty:
                continue

            df = dfs_flat[element_type]

            is_right_side = element_type in ("Feeds", "Products", "Junctions", "Misc Equipment")
            current_row = current_row_right if is_right_side else current_row_left
            start_col = 13 if is_right_side else 1

            start_table_row = current_row

            # 1. Write Title
            self._write_cell(
                worksheet, current_row, start_col, f"{element_type} Summary", style="title"
            )
            current_row += 2

            # 2. Write Table Content
            if element_type in ("Pumps", "Valves"):
                end_row, num_cols = self._write_transposed_table(
                    worksheet, df, current_row, start_col
                )
            else:
                end_row, num_cols = self._write_standard_table(
                    worksheet, df, current_row, start_col
                )

            # 3. Apply Styling & Borders
            self._apply_table_formatting(
                worksheet, start_table_row + 2, end_row, num_cols, start_col
            )

            # 4. Pipe stats block (max/min DP/DL and Velocity)
            if element_type == "Pipes":
                current_row = self._write_pipe_stats(worksheet, df, end_row + 2, start_col)
            else:
                current_row = end_row + 3
            if is_right_side:
                current_row_right = current_row
            else:
                current_row_left = current_row

        # Footer — auto-generated notice
        footer_row = max(current_row_left, current_row_right) + 1
        footer_text = (
            "This report is auto-generated from the KORF hydraulic model. "
            "Do not edit this document directly — any changes must be made in the source model "
            f"({source_name}) and the report regenerated."
        )
        footer_cell = worksheet.cell(row=footer_row, column=1, value=footer_text)
        footer_cell.font = self._styles["footer"]
        footer_cell.alignment = Alignment(wrap_text=False)
        worksheet.row_dimensions[footer_row].height = 30

        workbook.save(output_path)
        _logger.info("   Report saved | %s", output_path)
        return str(output_path)

    # =========================================================
    # INTERNAL WRITING HELPERS
    # =========================================================

    def _write_cell(
        self, ws: Any, row: int, col: int, value: Any, style: str | None = None
    ) -> None:
        """Helper to write a cell with an optional predefined style."""
        cell = ws.cell(row=row, column=col, value=value)
        if style and style in self._styles:
            cell.font = self._styles[style]
        return cell

    def _parse_headers(self, columns: list[str]) -> tuple[list[str], list[str]]:
        """Splits flat column names into (description, unit) tuples."""
        descriptions, units = [], []
        for col in columns:
            match = self._header_pattern.match(col)
            if match:
                descriptions.append(match.group(1))
                units.append(f"[{match.group(2)}]")
            else:
                descriptions.append(col)
                units.append("")
        return descriptions, units

    def _write_standard_table(
        self, ws: Any, df: pd.DataFrame, row: int, start_col: int = 1
    ) -> tuple[int, int]:
        """Writes a standard vertical table (Pipes, Compressors)."""
        descriptions, units = self._parse_headers(df.columns)

        # Write Two-Level Header
        ws.row_dimensions[row].height = 30
        for c_idx, (desc, unit) in enumerate(
            zip(descriptions, units, strict=True), start=start_col
        ):
            cell_desc = ws.cell(row=row, column=c_idx, value=desc)
            cell_desc.font = self._styles["header"]
            cell_desc.fill = self._styles["fill"]
            cell_desc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            cell_unit = ws.cell(row=row + 1, column=c_idx, value=unit)
            cell_unit.font = self._styles["unit"]
            cell_unit.fill = self._styles["fill"]
            cell_unit.alignment = Alignment(horizontal="center")

        # Write Data
        data_start_row = row + 2
        for r_idx, row_data in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=data_start_row
        ):
            for c_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = self._styles["data"]

        last_row = data_start_row + len(df) - 1
        self._apply_column_widths(ws, len(descriptions), start_col)

        return last_row, len(descriptions)

    def _write_transposed_table(
        self, ws: Any, df: pd.DataFrame, row: int, start_col: int = 1
    ) -> tuple[int, int]:
        """Writes a transposed table (Pumps, Valves)."""
        descriptions, units = self._parse_headers(df.columns)
        pump_names = df.iloc[:, 0].tolist()
        num_cols = 2 + len(pump_names)

        # Top Header Row (Parameter, Unit, Pump Names...)
        ws.row_dimensions[row].height = 30
        headers = ["Parameter", "Unit", *pump_names]
        for c_idx, val in enumerate(headers, start=start_col):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.font = self._styles["header"]
            cell.fill = self._styles["fill"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data Rows (One row per parameter)
        current_row = row + 1
        for p_idx in range(1, len(descriptions)):
            # Param Name
            cell_p = ws.cell(row=current_row, column=start_col, value=descriptions[p_idx])
            cell_p.font = self._styles["header"]
            cell_p.fill = self._styles["fill"]

            # Unit
            cell_u = ws.cell(row=current_row, column=start_col + 1, value=units[p_idx])
            cell_u.font = self._styles["unit"]

            # Values across pumps
            vals = df.iloc[:, p_idx].tolist()
            for v_idx, val in enumerate(vals, start=start_col + 2):
                cell_v = ws.cell(row=current_row, column=v_idx, value=val)
                cell_v.font = self._styles["data"]

            current_row += 1

        last_row = current_row - 1
        self._apply_column_widths(ws, num_cols, start_col)

        return last_row, num_cols

    def _apply_table_formatting(
        self, ws: Any, start_row: int, end_row: int, num_cols: int, start_col: int = 1
    ) -> None:
        """Applies internal borders and a thick outer border to a table block."""
        thin_s = self._styles["thin_side"]
        thick_s = self._styles["thick_side"]

        end_col = start_col + num_cols - 1
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)

                # Determine borders
                left = thick_s if c == start_col else thin_s
                right = thick_s if c == end_col else thin_s
                top = thick_s if r == start_row else thin_s
                bottom = thick_s if r == end_row else thin_s

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _apply_column_widths(self, ws: Any, num_cols: int, start_col: int = 1) -> None:
        """Sets fixed column widths: Col 1 = 25, Others = 15."""
        ws.column_dimensions[get_column_letter(start_col)].width = 25
        for c in range(start_col + 1, start_col + num_cols):
            ws.column_dimensions[get_column_letter(c)].width = 15

    def _get_model_title(self) -> str:
        """Fetches the model title from the first SYMBOL with TYPE='Text' and FSIZ=2."""
        from pykorf.elements import Symbol

        symbol_indices = {
            rec.index
            for rec in self.model._parser.records
            if rec.element_type == "SYMBOL" and rec.index is not None
        }
        for idx in symbol_indices:
            type_rec = self.model._parser.get("SYMBOL", idx, Symbol.TYPE)
            fsiz_rec = self.model._parser.get("SYMBOL", idx, Symbol.FSIZ)
            text_rec = self.model._parser.get("SYMBOL", idx, Symbol.TEXT)
            if not (type_rec and fsiz_rec and text_rec):
                continue
            try:
                if type_rec.values[0] == "Text" and int(fsiz_rec.values[0]) == 2:
                    return str(text_rec.values[0]) if text_rec.values else ""
            except (ValueError, TypeError, IndexError):
                continue
        return ""

    def _write_references_sheet(self, workbook: Any) -> None:
        """Creates the 'References & Design Basis' sheet as the first sheet (A4 Landscape 80%)."""
        ref_ws = workbook.create_sheet("References & Design Basis", 0)
        ref_ws.page_setup.paperSize = ref_ws.PAPERSIZE_A4
        ref_ws.page_setup.orientation = ref_ws.ORIENTATION_LANDSCAPE
        ref_ws.page_setup.scale = 80

        # Fixed column widths for all four content columns
        for col_idx, width in enumerate([50, 15, 15, 50], start=1):
            ref_ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = 1

        # Sheet title
        title_cell = ref_ws.cell(row=row, column=1, value="References & Design Basis")
        title_cell.font = self._styles["model_title"]
        ref_ws.row_dimensions[row].height = 30
        row += 2

        # ── Design Basis section ──────────────────────────────────────
        if self._basis and self._basis.strip():
            section_cell = ref_ws.cell(row=row, column=1, value="Design Basis")
            section_cell.font = self._styles["title"]
            row += 1

            for line in self._basis.splitlines():
                basis_cell = ref_ws.cell(row=row, column=1, value=line)
                basis_cell.font = Font(size=10)
                basis_cell.alignment = Alignment(wrap_text=False, horizontal="left")
                ref_ws.row_dimensions[row].height = 15
                row += 1
            row += 1  # blank separator

        # ── Remarks section ───────────────────────────────────────────
        if self._remarks and self._remarks.strip():
            section_cell = ref_ws.cell(row=row, column=1, value="Remarks")
            section_cell.font = self._styles["title"]
            row += 1

            for line in self._remarks.splitlines():
                cell = ref_ws.cell(row=row, column=1, value=line)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=False, horizontal="left")
                ref_ws.row_dimensions[row].height = 15
                row += 1
            row += 1  # blank separator

        # ── Hold Items section ────────────────────────────────────────
        if self._hold and self._hold.strip():
            section_cell = ref_ws.cell(row=row, column=1, value="Hold Items")
            section_cell.font = self._styles["title"]
            row += 1

            for line in self._hold.splitlines():
                cell = ref_ws.cell(row=row, column=1, value=line)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=False, horizontal="left")
                ref_ws.row_dimensions[row].height = 15
                row += 1
            row += 1  # blank separator

        # ── References table section ──────────────────────────────────
        if self._references:
            section_cell = ref_ws.cell(row=row, column=1, value="Reference Documents")
            section_cell.font = self._styles["title"]
            row += 1

            # Table headers
            headers = ["Name", "Category", "Link", "Description"]
            for c_idx, header in enumerate(headers, start=1):
                cell = ref_ws.cell(row=row, column=c_idx, value=header)
                cell.font = self._styles["header"]
                cell.fill = self._styles["fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ref_ws.row_dimensions[row].height = 20
            header_row = row
            row += 1

            # Table rows
            for ref in self._references:
                ref_ws.cell(row=row, column=1, value=ref.get("name", "")).font = Font(
                    bold=True, size=10
                )
                ref_ws.cell(row=row, column=2, value=ref.get("category", "")).font = Font(size=10)
                raw_link = ref.get("link", "")
                link_cell = ref_ws.cell(row=row, column=3, value="Open Link" if raw_link else "")
                if raw_link:
                    link_cell.hyperlink = raw_link
                    link_cell.font = Font(size=10, color="0563C1", underline="single")
                    link_cell.alignment = Alignment(wrap_text=False)
                ref_ws.cell(row=row, column=4, value=ref.get("description", "")).font = Font(
                    italic=True, size=10, color="555555"
                )
                row += 1

            # Outer border around the table
            self._apply_table_formatting(ref_ws, header_row, row - 1, len(headers), 1)

    def _write_pipe_stats(self, ws: Any, df: Any, row: int, start_col: int) -> int:
        """Writes a single Min - Max summary row below the pipe table."""
        import pandas as pd

        dpdl_col = next((c for c in df.columns if "DP / DL" in c and "Criteria" not in c), None)
        vel_col = next((c for c in df.columns if "Velocity" in c and "Criteria" not in c), None)
        if dpdl_col is None and vel_col is None:
            return row + 3

        col_names = list(df.columns)

        def _fmt(col: str) -> str | None:
            try:
                numeric = pd.to_numeric(df[col], errors="coerce")
                lo = numeric.min()
                hi = numeric.max()
                if pd.isna(lo) or pd.isna(hi):
                    return None
                lo_s = f"{lo:.4g}"
                hi_s = f"{hi:.4g}"
                return f"{lo_s} - {hi_s}"
            except Exception:
                return None

        # Label cell
        lc = ws.cell(row=row, column=start_col, value="Min - Max")
        lc.font = Font(bold=True, size=10)

        # DP/DL min-max
        if dpdl_col:
            val = _fmt(dpdl_col)
            c_idx = start_col + col_names.index(dpdl_col)
            ws.cell(row=row, column=c_idx, value=val).font = self._styles["data"]

        # Velocity min-max
        if vel_col:
            val = _fmt(vel_col)
            c_idx = start_col + col_names.index(vel_col)
            ws.cell(row=row, column=c_idx, value=val).font = self._styles["data"]

        return row + 3

    # =========================================================
    # ELEMENT EXTRACTORS
    # =========================================================

    def _extract_pipes(self) -> list[dict]:
        return [pipe.summary(export=True) for idx, pipe in self.model.pipes.items() if idx != 0]

    def _extract_pumps(self) -> list[dict]:
        return [pump.summary(export=True) for idx, pump in self.model.pumps.items() if idx != 0]

    def _extract_compressors(self) -> list[dict]:
        return [
            comp.summary(export=True) for idx, comp in self.model.compressors.items() if idx != 0
        ]

    def _extract_feeds(self) -> list[dict]:
        return [feed.summary(export=True) for idx, feed in self.model.feeds.items() if idx != 0]

    def _extract_products(self) -> list[dict]:
        return [prod.summary(export=True) for idx, prod in self.model.products.items() if idx != 0]

    def _extract_valves(self) -> list[dict]:
        return [valve.summary(export=True) for idx, valve in self.model.valves.items() if idx != 0]

    def _extract_heat_exchangers(self) -> list[dict]:
        return [hx.summary(export=True) for idx, hx in self.model.exchangers.items() if idx != 0]

    def _extract_junctions(self) -> list[dict]:
        """Extract junction data for export.

        Only includes junctions whose names do NOT start with "J".
        """
        return [
            junction.summary(export=True)
            for idx, junction in self.model.junctions.items()
            if idx != 0 and not junction.name.startswith("J")
        ]

    def _extract_misc(self) -> list[dict]:
        return [
            misc.summary(export=True) for idx, misc in self.model.misc_equipment.items() if idx != 0
        ]
