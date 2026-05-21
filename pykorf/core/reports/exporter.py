from __future__ import annotations

import logging
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from pykorf import Model
from pykorf.core.reports.formatting import (
    ReportStyles,
    apply_column_widths,
    apply_fail_format,
    apply_justified_format,
    apply_number_format,
    apply_table_borders,
    parse_headers,
    write_section_marker,
    write_transposed_header,
    write_two_level_headers,
)
from pykorf.core.reports.multi_case_summary import MultiCaseSummaryBuilder
from pykorf.core.reports.reporter import PykorfReporter, Reporter

_STYLES = ReportStyles()

_logger = logging.getLogger(__name__)

_RIGHT_SIDE_ELEMENTS = frozenset({"Feeds", "Products", "Junctions", "Misc Equipment"})
_TRANSPOSED_ELEMENTS = frozenset({"Pumps", "Valves"})
_STANDARD_ELEMENTS = frozenset({"Pipes", "Compressors", "Heat Exchangers", "Orifices"})


class ResultExporter:
    """A scalable exporter that generates formatted Excel reports from a Reporter.

    The Reporter provides the data (via PykorfReporter for KDF-based data,
    or KorfReporter for KORF Excel-based data), and this class handles all
    formatting, layout, and styling.

    Multi-case support: when the reporter returns multiple cases via
    ``generate_all_case_dataframes()``, the exporter creates:
    - A "Summary" envelope sheet (worst-case across all cases)
    - One sheet per case (named by case, e.g. "1 - Rated")
    - A "Validation" sheet with issues from all cases
    """

    def __init__(
        self,
        model: Model | None = None,
        reporter: Reporter | None = None,
        basis: str = "",
        remarks: str = "",
        hold: str = "",
        references: list[dict] | None = None,
        justifications: dict[str, str] | None = None,
    ):
        if reporter is not None:
            self.reporter = reporter
        elif model is not None:
            self.reporter = PykorfReporter(
                model=model,
                basis=basis,
                remarks=remarks,
                hold=hold,
                references=references,
                justifications=justifications,
            )
        else:
            raise ValueError("Either model or reporter must be provided")

    def export_to_excel(
        self,
        output_path: str,
        sheet_name: str = "Model Summary",
        elements: list[str] | None = None,
        pipe_columns: list[str] | None = None,
        template_path: str | None = None,
    ) -> str:
        """Generate a formatted Excel report.

        For multi-case reporters (KorfReporter), creates per-case sheets plus
        a Summary envelope sheet. For single-case reporters (PykorfReporter),
        creates the original single-sheet layout.
        """
        from pathlib import Path

        from openpyxl import load_workbook

        source_name = self.reporter.get_source_name()
        _logger.info("── Generate Report ── %s", source_name)

        all_cases = self.reporter.generate_all_case_dataframes()
        is_multi_case = len(all_cases) > 1

        if is_multi_case:
            non_empty_sections = set()
            for case_name, case_dfs in all_cases.items():
                for k, v in case_dfs.items():
                    if not v.empty:
                        non_empty_sections.add(k)
            _logger.info(
                "   Sections: %s (multi-case, %d cases)",
                ", ".join(sorted(non_empty_sections)) if non_empty_sections else "none",
                len(all_cases),
            )
        else:
            dfs_flat = self.reporter.generate_dataframes()
            non_empty = [k for k, v in dfs_flat.items() if not v.empty]
            _logger.info("   Sections: %s", ", ".join(non_empty) if non_empty else "none")

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        if (
            self.reporter.basis
            or self.reporter.remarks
            or self.reporter.hold
            or self.reporter.references
        ):
            self._write_references_sheet(workbook)

        if is_multi_case:
            self._export_multi_case(workbook, all_cases, source_name, elements, pipe_columns)
        else:
            template_ws = None
            if template_path and Path(template_path).exists():
                wb_template = load_workbook(template_path)
                template_ws = wb_template.active
                wb_template.close()

            self._export_single_case(
                workbook,
                dfs_flat,
                source_name,
                sheet_name,
                elements,
                template_ws,
                pipe_columns,
            )

        self._write_validation_sheet(workbook)
        workbook.save(output_path)
        _logger.info("   Report saved | %s", output_path)
        return str(output_path)

    def _export_single_case(
        self,
        workbook: openpyxl.Workbook,
        dfs_flat: dict[str, pd.DataFrame],
        source_name: str,
        sheet_name: str,
        elements: list[str] | None,
        template_ws: Any | None = None,
        pipe_columns: list[str] | None = None,
    ) -> None:
        """Write a single-sheet report (original layout)."""
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1

        current_row = self._write_model_title(worksheet)

        worksheet.cell(row=current_row, column=1, value=f"Source File: {source_name}").font = _STYLES.header

        case_names = self.reporter.get_case_names()
        if case_names:
            worksheet.cell(
                row=current_row + 1, column=1, value=f"Cases: {'; '.join(case_names)}"
            ).font = _STYLES.header

        if pipe_columns and "Pipes" in dfs_flat and not dfs_flat["Pipes"].empty:
            dfs_flat = dict(dfs_flat)
            dfs_flat["Pipes"] = self._filter_pipe_columns(dfs_flat["Pipes"], pipe_columns)

        current_row_left = 5
        current_row_right = 5
        max_left_col = 1

        element_keys = elements if elements is not None else list(dfs_flat.keys())

        for element_type in element_keys:
            if element_type not in dfs_flat or dfs_flat[element_type].empty:
                continue
            if element_type in _TRANSPOSED_ELEMENTS:
                continue
            if element_type in _RIGHT_SIDE_ELEMENTS:
                continue
            df = dfs_flat[element_type]
            num_cols = len(df.columns)
            end_col = 1 + num_cols - 1
            if end_col > max_left_col:
                max_left_col = end_col

        for element_type in element_keys:
            if element_type not in dfs_flat or dfs_flat[element_type].empty:
                continue

            df = dfs_flat[element_type]
            is_right_side = element_type in _RIGHT_SIDE_ELEMENTS
            current_row = current_row_right if is_right_side else current_row_left
            start_col = max_left_col + 2 if is_right_side else 1

            start_table_row = current_row

            self._write_cell(
                worksheet, current_row, start_col, f"{element_type} Summary", style="title"
            )
            current_row += 2

            if element_type in _TRANSPOSED_ELEMENTS:
                end_row, num_cols = self._write_transposed_table(
                    worksheet, df, current_row, start_col
                )
            else:
                end_row, num_cols = self._write_standard_table(
                    worksheet, df, current_row, start_col
                )

            apply_table_borders(worksheet, start_table_row + 2, end_row, num_cols, start_col)

            if element_type == "Pipes":
                current_row = self._write_pipe_stats(worksheet, df, end_row + 2, start_col)
            else:
                current_row = end_row + 3

            if is_right_side:
                current_row_right = current_row
            else:
                current_row_left = current_row

        footer_row = max(current_row_left, current_row_right) + 1
        footer_text = (
            "This report is auto-generated from the KORF hydraulic model. "
            "Do not edit this document directly — any changes must be made in the source model "
            f"({source_name}) and the report regenerated."
        )
        footer_cell = worksheet.cell(row=footer_row, column=1, value=footer_text)
        footer_cell.font = _STYLES.footer
        footer_cell.alignment = Alignment(wrap_text=False)
        worksheet.row_dimensions[footer_row].height = 30

    def _export_multi_case(
        self,
        workbook: openpyxl.Workbook,
        all_cases: dict[str, dict[str, pd.DataFrame]],
        source_name: str,
        elements: list[str] | None,
        pipe_columns: list[str] | None = None,
    ) -> None:
        """Write per-case sheets plus a Summary envelope sheet for multi-case reports."""
        case_names = list(all_cases.keys())

        first_case_dfs = list(all_cases.values())[0]
        element_keys = elements if elements is not None else list(first_case_dfs.keys())

        summary_ws = workbook.create_sheet("Summary")
        summary_ws.page_setup.paperSize = summary_ws.PAPERSIZE_A3
        summary_ws.page_setup.orientation = summary_ws.ORIENTATION_LANDSCAPE
        summary_ws.page_setup.fitToPage = True
        summary_ws.page_setup.fitToWidth = 1
        summary_ws.page_setup.fitToHeight = 1

        current_row = self._write_model_title(summary_ws)

        if isinstance(
            self.reporter,
            __import__("pykorf.core.reports.korf_reporter", fromlist=["KorfReporter"]).KorfReporter,
        ):
            case_data = self.reporter._get_case_data()
            builder = MultiCaseSummaryBuilder(
                case_data=case_data,
                model=self.reporter.model,
                reporter=self.reporter,
                justifications=self.reporter._justifications,
            )
            builder.write_summary_sheet(
                summary_ws,
                source_name,
                pipe_stats_handler=self._write_pipe_stats,
                pipe_columns=pipe_columns,
            )
        else:
            envelope_dfs = self._build_envelope_dataframes(all_cases, element_keys)
            if pipe_columns and "Pipes" in envelope_dfs and not envelope_dfs["Pipes"].empty:
                envelope_dfs = dict(envelope_dfs)
                envelope_dfs["Pipes"] = self._filter_pipe_columns(
                    envelope_dfs["Pipes"], pipe_columns
                )
            self._write_case_sheet_content(summary_ws, envelope_dfs, source_name, element_keys)
            case_label = f"Cases: {'; '.join(case_names)}"
            summary_ws.cell(row=3, column=1, value=case_label).font = _STYLES.header

        for case_name in case_names:
            case_dfs = all_cases[case_name]
            ws = self._write_case_sheet(
                workbook, case_name, case_dfs, source_name, element_keys, pipe_columns
            )
            ws.cell(row=3, column=1, value=f"Case: {case_name}").font = _STYLES.header

    def _write_case_sheet_content(
        self,
        ws: Any,
        dfs: dict[str, pd.DataFrame],
        source_name: str,
        elements: list[str] | None,
    ) -> None:
        """Write element tables content to a worksheet (without header/footer setup)."""
        element_keys = elements if elements is not None else list(dfs.keys())

        current_row_left = 5
        current_row_right = 5
        max_left_col = 1

        for element_type in element_keys:
            if element_type not in dfs or dfs[element_type].empty:
                continue
            if element_type in _TRANSPOSED_ELEMENTS:
                continue
            if element_type in _RIGHT_SIDE_ELEMENTS:
                continue
            df = dfs[element_type]
            num_cols = len(df.columns)
            end_col = 1 + num_cols - 1
            if end_col > max_left_col:
                max_left_col = end_col

        for element_type in element_keys:
            if element_type not in dfs or dfs[element_type].empty:
                continue

            df = dfs[element_type]
            is_right_side = element_type in _RIGHT_SIDE_ELEMENTS
            current_row = current_row_right if is_right_side else current_row_left
            start_col = max_left_col + 2 if is_right_side else 1

            start_table_row = current_row

            self._write_cell(ws, current_row, start_col, f"{element_type} Summary", style="title")
            current_row += 2

            if element_type in _TRANSPOSED_ELEMENTS:
                end_row, num_cols = self._write_transposed_table(ws, df, current_row, start_col)
            else:
                end_row, num_cols = self._write_standard_table(ws, df, current_row, start_col)

            apply_table_borders(ws, start_table_row + 2, end_row, num_cols, start_col)

            if element_type == "Pipes":
                current_row = self._write_pipe_stats(ws, df, end_row + 2, start_col)
            else:
                current_row = end_row + 3

            if is_right_side:
                current_row_right = current_row
            else:
                current_row_left = current_row

        footer_row = max(current_row_left, current_row_right) + 1
        footer_text = (
            "This report is auto-generated from the KORF hydraulic model. "
            "Do not edit this document directly — any changes must be made in the source model "
            f"({source_name}) and the report regenerated."
        )
        footer_cell = ws.cell(row=footer_row, column=1, value=footer_text)
        footer_cell.font = _STYLES.footer
        footer_cell.alignment = Alignment(wrap_text=False)
        ws.row_dimensions[footer_row].height = 30

    def _write_case_sheet(
        self,
        workbook: openpyxl.Workbook,
        sheet_name: str,
        dfs: dict[str, pd.DataFrame],
        source_name: str,
        elements: list[str] | None,
        pipe_columns: list[str] | None = None,
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Write a single worksheet with element tables from one case's DataFrames."""
        ws = workbook.create_sheet(sheet_name)
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1

        current_row = self._write_model_title(ws)

        ws.cell(row=current_row, column=1, value=f"Source File: {source_name}").font = _STYLES.header

        if pipe_columns and "Pipes" in dfs and not dfs["Pipes"].empty:
            dfs = dict(dfs)
            dfs["Pipes"] = self._filter_pipe_columns(dfs["Pipes"], pipe_columns)

        element_keys = elements if elements is not None else list(dfs.keys())

        current_row_left = 5
        current_row_right = 5
        max_left_col = 1

        for element_type in element_keys:
            if element_type not in dfs or dfs[element_type].empty:
                continue
            if element_type in _TRANSPOSED_ELEMENTS:
                continue
            if element_type in _RIGHT_SIDE_ELEMENTS:
                continue
            df = dfs[element_type]
            num_cols = len(df.columns)
            end_col = 1 + num_cols - 1
            if end_col > max_left_col:
                max_left_col = end_col

        for element_type in element_keys:
            if element_type not in dfs or dfs[element_type].empty:
                continue

            df = dfs[element_type]
            is_right_side = element_type in _RIGHT_SIDE_ELEMENTS
            current_row = current_row_right if is_right_side else current_row_left
            start_col = max_left_col + 2 if is_right_side else 1

            start_table_row = current_row

            self._write_cell(ws, current_row, start_col, f"{element_type} Summary", style="title")
            current_row += 2

            if element_type in _TRANSPOSED_ELEMENTS:
                end_row, num_cols = self._write_transposed_table(ws, df, current_row, start_col)
            else:
                end_row, num_cols = self._write_standard_table(ws, df, current_row, start_col)

            apply_table_borders(ws, start_table_row + 2, end_row, num_cols, start_col)

            if element_type == "Pipes":
                current_row = self._write_pipe_stats(ws, df, end_row + 2, start_col)
            else:
                current_row = end_row + 3

            if is_right_side:
                current_row_right = current_row
            else:
                current_row_left = current_row

        footer_row = max(current_row_left, current_row_right) + 1
        footer_text = (
            "This report is auto-generated from the KORF hydraulic model. "
            "Do not edit this document directly — any changes must be made in the source model "
            f"({source_name}) and the report regenerated."
        )
        footer_cell = ws.cell(row=footer_row, column=1, value=footer_text)
        footer_cell.font = _STYLES.footer
        footer_cell.alignment = Alignment(wrap_text=False)
        ws.row_dimensions[footer_row].height = 30

        return ws

    def _resolve_pipe_columns(self, df: pd.DataFrame, keys: list[str]) -> list[str]:
        """Resolve partial column keys to actual DataFrame column names.

        Does exact match first, then prefix match (e.g. "Velocity" → "Velocity [m/s]").
        """
        available = set(df.columns)
        resolved: list[str] = []
        for key in keys:
            if key in available:
                resolved.append(key)
            else:
                for col in df.columns:
                    if col.startswith(key + " [") or col == key:
                        resolved.append(col)
                        break
        return resolved

    def _filter_pipe_columns(self, df: pd.DataFrame, pipe_columns: list[str]) -> pd.DataFrame:
        """Filter pipe DataFrame to only include requested columns.

        Always includes 'Pipe Name' and 'Criteria Check' regardless of selection.
        Preserves the original DataFrame column order.
        """
        always_keys = {"Pipe Name", "Governing Case", "Criteria Check"}
        resolved = self._resolve_pipe_columns(df, pipe_columns)
        always_resolved = self._resolve_pipe_columns(df, list(always_keys))
        selected_set = set(resolved) | set(always_resolved)
        ordered = [c for c in df.columns if c in selected_set]
        if not ordered:
            return df
        return df[ordered]

    def _build_envelope_dataframes(
        self,
        all_cases: dict[str, dict[str, pd.DataFrame]],
        element_keys: list[str],
    ) -> dict[str, pd.DataFrame]:
        """Build worst-case envelope DataFrames across all cases.

        For Pipes: takes the row with the worst criteria result (FAIL > JUSTIFIED > PASS).
        For other elements: takes the row from the first case.
        Transposed elements (Pumps, Valves): takes data from the first case.
        """
        case_names = list(all_cases.keys())
        if not case_names:
            return {}

        first_case = all_cases[case_names[0]]
        envelope: dict[str, pd.DataFrame] = {}

        for element_type in element_keys:
            if element_type not in first_case or first_case[element_type].empty:
                continue

            if element_type == "Pipes":
                envelope[element_type] = self._build_pipe_envelope(all_cases, element_type)
            elif element_type in _TRANSPOSED_ELEMENTS:
                envelope[element_type] = self._build_transposed_envelope(all_cases, element_type)
            else:
                envelope[element_type] = first_case[element_type].copy()

        return envelope

    def _build_pipe_envelope(
        self,
        all_cases: dict[str, dict[str, pd.DataFrame]],
        element_type: str,
    ) -> pd.DataFrame:
        """Build worst-case pipe envelope: for each pipe, pick the case row with worst criteria."""
        case_names = list(all_cases.keys())
        base_df = all_cases[case_names[0]][element_type].copy()

        if len(case_names) <= 1:
            return base_df

        name_col = base_df.columns[0] if len(base_df.columns) > 0 else None
        if name_col is None:
            return base_df

        def _worst_criteria(*values: str) -> str:
            priority = {"FAIL": 0, "JUSTIFIED": 1, "PASS": 2, "": 3}
            return min(values, key=lambda v: priority.get(v, 3))

        result_rows = []
        for _, base_row in base_df.iterrows():
            pipe_name = base_row[name_col]
            candidate_rows = []

            for cn in case_names:
                case_df = all_cases[cn].get(element_type, pd.DataFrame())
                if case_df.empty:
                    continue
                match = case_df[case_df[name_col] == pipe_name]
                if match.empty:
                    continue
                candidate_rows.append(match.iloc[0])

            if not candidate_rows:
                result_rows.append(base_row)
                continue

            if "Criteria Check" not in base_df.columns:
                result_rows.append(candidate_rows[0])
                continue

            worst_row = candidate_rows[0]
            worst_val = str(worst_row.get("Criteria Check", ""))
            for row in candidate_rows[1:]:
                val = str(row.get("Criteria Check", ""))
                if _worst_criteria(val, worst_val) != worst_val:
                    worst_row = row
                    worst_val = val

            result_rows.append(worst_row)

        if result_rows:
            return pd.DataFrame(result_rows).reset_index(drop=True)
        return base_df

    def _build_transposed_envelope(
        self,
        all_cases: dict[str, dict[str, pd.DataFrame]],
        element_type: str,
    ) -> pd.DataFrame:
        """Build envelope for transposed elements (Pumps, Valves): use first case data."""
        case_names = list(all_cases.keys())
        return all_cases[case_names[0]][element_type].copy()

    # =========================================================
    # INTERNAL WRITING HELPERS
    # =========================================================

    def _write_model_title(self, ws: Any) -> int:
        """Write the model title to A1 of *ws*. Returns 2 if title was written, else 1."""
        title = self.reporter.get_model_title()
        if title:
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = _STYLES.model_title
            ws.row_dimensions[1].height = 28
            return 2
        return 1

    def _write_cell(
        self, ws: Any, row: int, col: int, value: Any, style: str | None = None
    ) -> None:
        """Helper to write a cell with an optional predefined style."""
        cell = ws.cell(row=row, column=col, value=value)
        if style:
            style_font = getattr(_STYLES, style, None)
            if style_font is not None:
                cell.font = style_font
        return cell

    def _write_standard_table(
        self, ws: Any, df: pd.DataFrame, row: int, start_col: int = 1
    ) -> tuple[int, int]:
        """Writes a standard vertical table (Pipes, Compressors)."""
        descriptions, units = parse_headers(df.columns)

        write_two_level_headers(ws, row, start_col, descriptions, units)

        # Write Data
        data_start_row = row + 2
        criteria_col_idx = (
            start_col + len(descriptions) - 1 if "Criteria Check" in descriptions else None
        )
        for r_idx, row_data in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=data_start_row
        ):
            for c_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = _STYLES.data
                if c_idx != start_col:
                    cell.alignment = _STYLES.data_center_align
                if c_idx > start_col:
                    col_name = df.columns[c_idx - start_col]
                    apply_number_format(cell, col_name)
                if criteria_col_idx is not None and c_idx == criteria_col_idx:
                    if val == "FAIL":
                        apply_fail_format(cell)
                    elif val == "JUSTIFIED":
                        apply_justified_format(cell)

        last_row = data_start_row + len(df) - 1
        apply_column_widths(ws, len(descriptions), start_col)

        # Auto-fit "Line Number" column if present
        line_num_idx = None
        for i, col in enumerate(df.columns):
            if col == "Line Number":
                line_num_idx = start_col + i
                break
        if line_num_idx is not None:
            max_len = max(len(str(v)) for v in df["Line Number"].fillna("")) if len(df) else 0
            ws.column_dimensions[get_column_letter(line_num_idx)].width = min(
                max(max_len + 2, 8), 40
            )

        return last_row, len(descriptions)

    def _write_transposed_table(
        self, ws: Any, df: pd.DataFrame, row: int, start_col: int = 1
    ) -> tuple[int, int]:
        """Writes a transposed table (Pumps, Valves).

        Detects section markers (rows where parameter name is empty and value is non-empty)
        and renders them as styled separator rows with section headings.
        """
        descriptions, units = parse_headers(df.columns)
        pump_names = df.iloc[:, 0].tolist()
        num_cols = 2 + len(pump_names)

        write_transposed_header(ws, row, start_col, ["Parameter", "Unit", *pump_names])

        # Data Rows (One row per parameter)
        current_row = row + 1
        for p_idx in range(1, len(descriptions)):
            param_name = descriptions[p_idx]
            unit_text = units[p_idx]

            row_values = df.iloc[:, p_idx].tolist()
            is_section_marker = (
                param_name
                and unit_text == ""
                and len(row_values) > 0
                and (
                    row_values[0] == param_name
                    or (param_name.startswith("Section_") and row_values[0] == param_name[8:])
                )
            )

            if is_section_marker:
                section_name = param_name[8:] if param_name.startswith("Section_") else param_name
                write_section_marker(ws, current_row, start_col, num_cols, section_name)
            else:
                cell_p = ws.cell(row=current_row, column=start_col, value=param_name)
                cell_p.font = _STYLES.header
                cell_p.fill = _STYLES.header_fill

                cell_u = ws.cell(row=current_row, column=start_col + 1, value=unit_text)
                cell_u.font = _STYLES.unit

                vals = df.iloc[:, p_idx].tolist()
                for v_idx, val in enumerate(vals, start=start_col + 2):
                    cell_v = ws.cell(row=current_row, column=v_idx, value=val)
                    cell_v.font = _STYLES.data
                    cell_v.alignment = _STYLES.data_center_align
                    apply_number_format(cell_v, param_name)

            current_row += 1

        last_row = current_row - 1
        apply_column_widths(ws, num_cols, start_col)

        return last_row, num_cols

    def _write_references_sheet(self, workbook: Any) -> None:
        """Creates the 'References & Design Basis' sheet as the first sheet (A4 Landscape 80%)."""
        ref_ws = workbook.create_sheet("References & Design Basis", 0)
        assert ref_ws is not None, "Failed to create References & Design Basis sheet"
        ref_ws.page_setup.paperSize = ref_ws.PAPERSIZE_A4
        ref_ws.page_setup.orientation = ref_ws.ORIENTATION_LANDSCAPE
        ref_ws.page_setup.scale = 75

        # Fixed column widths for Name, Category, Link; Description auto-fits below
        for col_idx, width in enumerate([50, 15, 15], start=1):
            ref_ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = self._write_model_title(ref_ws)

        # Sheet title
        title_cell = ref_ws.cell(row=row, column=1, value="References & Design Basis")
        title_cell.font = _STYLES.title
        ref_ws.row_dimensions[row].height = 30
        row += 2

        # ── Design Basis section ──────────────────────────────────────
        if self.reporter.basis and self.reporter.basis.strip():
            section_cell = ref_ws.cell(row=row, column=1, value="Design Basis")
            section_cell.font = _STYLES.title
            row += 1

            for line in self.reporter.basis.splitlines():
                basis_cell = ref_ws.cell(row=row, column=1, value=line)
                basis_cell.font = Font(size=10)
                basis_cell.alignment = Alignment(wrap_text=False, horizontal="left")
                ref_ws.row_dimensions[row].height = 15
                row += 1
            row += 1  # blank separator

        # ── Remarks section ───────────────────────────────────────────
        if self.reporter.remarks and self.reporter.remarks.strip():
            section_cell = ref_ws.cell(row=row, column=1, value="Remarks")
            section_cell.font = _STYLES.title
            row += 1

            for line in self.reporter.remarks.splitlines():
                cell = ref_ws.cell(row=row, column=1, value=line)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=False, horizontal="left")
                ref_ws.row_dimensions[row].height = 15
                row += 1
            row += 1  # blank separator

        # ── Hold Items section ────────────────────────────────────────
        if self.reporter.hold and self.reporter.hold.strip():
            section_cell = ref_ws.cell(row=row, column=1, value="Hold Items")
            section_cell.font = _STYLES.title
            row += 1

            for line in self.reporter.hold.splitlines():
                cell = ref_ws.cell(row=row, column=1, value=line)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=False, horizontal="left")
                ref_ws.row_dimensions[row].height = 15
                row += 1
            row += 1  # blank separator

        # ── References table section ──────────────────────────────────
        if self.reporter.references:
            section_cell = ref_ws.cell(row=row, column=1, value="Reference Documents")
            section_cell.font = _STYLES.title
            row += 1

            # Table headers
            headers = ["Name", "Category", "Link", "Description"]
            for c_idx, header in enumerate(headers, start=1):
                cell = ref_ws.cell(row=row, column=c_idx, value=header)
                cell.font = _STYLES.header
                cell.fill = _STYLES.header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ref_ws.row_dimensions[row].height = 20
            header_row = row
            row += 1

            # Table rows
            for ref in self.reporter.references:
                ref_ws.cell(row=row, column=1, value=ref.get("name", "")).font = Font(
                    bold=True, size=10
                )
                ref_ws.cell(row=row, column=2, value=ref.get("category", "")).font = Font(size=10)
                raw_link = ref.get("link", "")
                link_cell = ref_ws.cell(row=row, column=3, value="Open Link" if raw_link else "")
                if raw_link:
                    link_cell.hyperlink = raw_link
                    link_cell.font = Font(size=10, color="0563C1", underline="single")
                    link_cell.alignment = Alignment(wrap_text=False)
                ref_ws.cell(row=row, column=4, value=ref.get("description", "")).font = Font(
                    italic=True, size=10, color="555555"
                )
                row += 1

            # Auto-fit Description column (column 4)
            desc_col_letter = get_column_letter(4)
            desc_max = max(
                (len(str(ref.get("description", ""))) for ref in self.reporter.references),
                default=0,
            )
            ref_ws.column_dimensions[desc_col_letter].width = min(max(desc_max + 2, 10), 80)

            # Outer border around the table
            apply_table_borders(ref_ws, header_row, row - 1, len(headers), 1)

    def _write_validation_sheet(self, workbook: Any) -> None:
        """Create a 'Validation' sheet with model-level and connectivity issues.

        Only added if there are actual issues. A4 Landscape, 90% scale.
        """
        df = self.reporter.get_validation_dataframe()
        if df.empty:
            return

        val_ws = workbook.create_sheet("Validation")
        assert val_ws is not None, "Failed to create Validation sheet"
        val_ws.page_setup.paperSize = val_ws.PAPERSIZE_A4
        val_ws.page_setup.orientation = val_ws.ORIENTATION_LANDSCAPE
        val_ws.page_setup.scale = 90

        for col_idx, width in enumerate([12, 18, 15, 80], start=1):
            val_ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = self._write_model_title(val_ws)

        # Sheet title
        title_cell = val_ws.cell(row=row, column=1, value="Validation Results")
        title_cell.font = _STYLES.title
        val_ws.row_dimensions[row].height = 28
        row += 1

        # Issue count summary
        error_count = (df["Severity"] == "Error").sum()
        warning_count = (df["Severity"] == "Warning").sum()
        info_count = (df["Severity"] == "Info").sum()
        justified_count = (df["Severity"] == "Justified").sum()
        summary_parts = [f"{len(df)} issue(s)"]
        if error_count:
            summary_parts.append(f"{error_count} error(s)")
        if warning_count:
            summary_parts.append(f"{warning_count} warning(s)")
        if info_count:
            summary_parts.append(f"{info_count} info")
        if justified_count:
            summary_parts.append(f"{justified_count} justified")
        summary = " — ".join(summary_parts)
        val_ws.cell(row=row, column=1, value=summary).font = Font(
            italic=True, size=10, color="555555"
        )
        row += 2

        # Table headers
        headers = list(df.columns)
        severity_colors: dict[str, str] = {
            "Error": "9C0006",
            "Warning": "9C5700",
            "Info": "003366",
            "Justified": "003366",
        }
        severity_fills: dict[str, str] = {
            "Error": "FFC7CE",
            "Warning": "FFEB9C",
            "Info": "D9E2F3",
            "Justified": "D9E2F3",
        }
        for c_idx, header in enumerate(headers, start=1):
            cell = val_ws.cell(row=row, column=c_idx, value=header)
            cell.font = _STYLES.header
            cell.fill = _STYLES.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        val_ws.row_dimensions[row].height = 20
        header_row = row
        row += 1

        # Table rows
        for _, r in df.iterrows():
            severity = str(r.get("Severity", "")) if r.get("Severity", "") is not None else ""
            color = severity_colors.get(severity, "555555")
            fill = severity_fills.get(severity, "F2F2F2")

            val_ws.cell(
                row=row,
                column=1,
                value=str(r.get("Severity", "")) if r.get("Severity", "") is not None else "",
            ).font = Font(bold=True, size=10, color=color)
            val_ws.cell(row=row, column=1).fill = PatternFill(
                start_color=fill, end_color=fill, fill_type="solid"
            )
            val_ws.cell(row=row, column=2, value=r.get("Category", "")).font = Font(size=10)
            val_ws.cell(row=row, column=3, value=r.get("Element", "")).font = Font(
                bold=True, size=10
            )
            msg_cell = val_ws.cell(row=row, column=4, value=r.get("Message", ""))
            msg_cell.font = Font(size=10)
            msg_cell.alignment = Alignment(wrap_text=True, vertical="top")
            val_ws.row_dimensions[row].height = 15
            row += 1

        # Outer border
        apply_table_borders(val_ws, header_row, row - 1, len(headers), 1)

    def _write_pipe_stats(self, ws: Any, df: Any, row: int, start_col: int) -> int:
        """Writes a Min-Max summary row and an overall Criteria Check row below the pipe table."""
        dpdl_col = next((c for c in df.columns if "DP / DL" in c and "Criteria" not in c), None)
        vel_col = next((c for c in df.columns if "Velocity" in c and "Criteria" not in c), None)
        rhov2_col = next((c for c in df.columns if "ρV² calc" in c), None)  # noqa: RUF001
        criteria_col = next((c for c in df.columns if c == "Criteria Check"), None)

        if dpdl_col is None and vel_col is None and rhov2_col is None and criteria_col is None:
            return row + 3

        col_names = list(df.columns)
        criteria_col_idx = start_col + col_names.index(criteria_col) if criteria_col else None

        def _fmt(col: str, thousand_comma: bool = False) -> str | None:
            try:
                numeric = pd.to_numeric(df[col], errors="coerce")
                lo, hi = numeric.min(), numeric.max()
                if pd.isna(lo) or pd.isna(hi):
                    return None
                if thousand_comma:
                    return f"{lo:,.0f} - {hi:,.0f}"
                return f"{lo:.4g} - {hi:.4g}"
            except Exception:
                return None

        # Collect column indices that have data in Min-Max row
        data_col_indices = []
        if dpdl_col:
            data_col_indices.append(start_col + col_names.index(dpdl_col))
        if vel_col:
            data_col_indices.append(start_col + col_names.index(vel_col))
        if rhov2_col:
            data_col_indices.append(start_col + col_names.index(rhov2_col))
        if criteria_col:
            data_col_indices.append(criteria_col_idx)

        min_max_cols = [start_col, *data_col_indices]
        start_col_idx = min(min_max_cols)
        end_col_idx = max(min_max_cols)
        thin_s = _STYLES.thin_side
        thick_s = _STYLES.thick_side

        # -- Row 1: Min - Max ----------------------------------------------------
        lc = ws.cell(row=row, column=start_col, value="Min - Max")
        lc.font = Font(bold=True, size=10)

        if dpdl_col:
            c_idx = start_col + col_names.index(dpdl_col)
            cell = ws.cell(row=row, column=c_idx, value=_fmt(dpdl_col))
            cell.font = _STYLES.data
            cell.alignment = Alignment(horizontal="center")

        if vel_col:
            c_idx = start_col + col_names.index(vel_col)
            cell = ws.cell(row=row, column=c_idx, value=_fmt(vel_col))
            cell.font = _STYLES.data
            cell.alignment = Alignment(horizontal="center")

        if rhov2_col:
            c_idx = start_col + col_names.index(rhov2_col)
            cell = ws.cell(row=row, column=c_idx, value=_fmt(rhov2_col, thousand_comma=True))
            cell.font = _STYLES.data
            cell.alignment = Alignment(horizontal="center")

        # Apply borders to Min-Max row
        for c_idx in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=row, column=c_idx)
            left = thick_s if c_idx == start_col_idx else thin_s
            right = thick_s if c_idx == end_col_idx else thin_s
            top = thick_s
            bottom = thin_s
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            if c_idx not in min_max_cols:
                cell.alignment = Alignment(horizontal="center")

        # -- Row 2: Overall Criteria Check (OR logic - any FAIL -> FAIL) --------
        if criteria_col:
            has_fail = (df[criteria_col] == "FAIL").any()
            has_justified = (df[criteria_col] == "JUSTIFIED").any()

            if has_fail:
                overall = "FAIL"
                color = "9C0006"
                fill = "FFC7CE"
            elif has_justified:
                overall = "PASS with Justification"
                color = "003366"
                fill = "D9E2F3"
            else:
                overall = "PASS"
                color = "276221"
                fill = "C6EFCE"

            label_cell = ws.cell(row=row + 1, column=start_col, value="Overall Criteria")
            label_cell.font = Font(bold=True, size=10)

            c_idx = start_col + col_names.index(criteria_col)
            result_cell = ws.cell(row=row + 1, column=c_idx, value=overall)
            result_cell.font = Font(
                bold=True,
                size=10,
                color=color,
            )
            result_cell.fill = PatternFill(
                start_color=fill,
                end_color=fill,
                fill_type="solid",
            )
            result_cell.alignment = Alignment(horizontal="center")

            # Apply borders to Overall Criteria row
            for c_idx in range(start_col_idx, end_col_idx + 1):
                cell = ws.cell(row=row + 1, column=c_idx)
                left = thick_s if c_idx == start_col_idx else thin_s
                right = thick_s if c_idx == end_col_idx else thin_s
                top = thin_s
                bottom = thick_s
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                if c_idx != start_col and c_idx != criteria_col_idx:
                    cell.alignment = Alignment(horizontal="center")

        return row + 3
