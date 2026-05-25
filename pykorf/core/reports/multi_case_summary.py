"""Multi-case Summary sheet builder for KORF Excel reports."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING, Any

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from pykorf import Model
from pykorf.core.reports.formatting import (
    ReportStyles,
    apply_fail_format,
    apply_justified_format,
    apply_number_format,
    apply_table_borders,
    parse_header_unit,
    parse_headers,
    write_section_marker,
    write_transposed_header,
    write_two_level_headers,
)
from pykorf.core.reports.korf_parser import CaseInfo, KorfCaseData, PipeData, PumpData
from pykorf.core.reports.unit_converter import UnitConverter

if TYPE_CHECKING:
    from pykorf.core.reports.korf_reporter import KorfReporter

_STYLES = ReportStyles()

_logger = logging.getLogger(__name__)

_VALVE_PARAMETERS = [
    ("Flowrate", "kg/h"),
    ("Inlet Pressure", "barg"),
    ("Differential Pressure", "bar"),
]


class MultiCaseSummaryBuilder:
    """Builds the Summary sheet for multi-case KORF Excel reports.

    This class generates element summary data with governing case determination:
    - Pipes: Governing case based on maximum velocity
    - Pumps: Governing case is the first case
    - Valves: Per-element tables with all case values

    Excluded from Summary: Feeds, Products, Orifices

    Placeholder TODO methods for: Compressors, Heat Exchangers, Misc Equipment,
    Feeds, Products, Orifices (to be implemented later).
    """

    def __init__(
        self,
        case_data: dict[CaseInfo, KorfCaseData],
        model: Model,
        reporter: KorfReporter,
        justifications: dict[str, str] | None = None,
    ):
        self._case_data = case_data
        self._model = model
        self._reporter = reporter
        self._justifications = justifications or {}
        self._case_names = self._get_case_names()

    def _get_case_names(self) -> list[str]:
        """Get case names in order (e.g., ['Rated', 'Case2'])."""
        sorted_cases = sorted(self._case_data.keys(), key=lambda c: int(c.number))
        return [c.name for c in sorted_cases]

    def _get_full_case_names(self) -> list[str]:
        """Get full case names (e.g., ['1 - Rated', '2 - Case2'])."""
        sorted_cases = sorted(self._case_data.keys(), key=lambda c: int(c.number))
        return [f"{c.number} - {c.name}" for c in sorted_cases]

    def write_summary_sheet(
        self,
        ws: Worksheet,
        source_name: str,
        pipe_stats_handler: Any | None = None,
        pipe_columns: list[str] | None = None,
    ) -> None:
        """Write the complete Summary sheet content.

        Args:
            ws: Worksheet to write to
            source_name: Source file name for display
            pipe_stats_handler: Optional callable to write pipe stats (Min-Max, Overall Criteria)
            pipe_columns: Optional subset of pipe columns to include.
        """
        current_row = 2

        ws.cell(
            row=current_row, column=1, value=f"Source File: {source_name}"
        ).font = _STYLES.header
        current_row += 1

        case_label = f"Cases: {'; '.join(self._get_full_case_names())}"
        ws.cell(row=current_row, column=1, value=case_label).font = _STYLES.header
        current_row += 2

        pipe_df = self.build_pipe_summary_df()
        if not pipe_df.empty:
            if pipe_columns:
                pipe_df = self._filter_pipe_columns(pipe_df, pipe_columns)
            ws.cell(row=current_row, column=1, value="Pipes Summary").font = _STYLES.title
            current_row += 2
            end_row = self._write_pipe_table(ws, pipe_df, current_row)
            if pipe_stats_handler:
                current_row = pipe_stats_handler(ws, pipe_df, end_row + 2, 1)
            else:
                current_row = end_row + 3

        pump_df = self.build_pump_summary_df()
        if not pump_df.empty:
            ws.cell(row=current_row, column=1, value="Pumps Summary").font = _STYLES.title
            current_row += 2
            end_row = self._write_pump_table(ws, pump_df, current_row)
            current_row = end_row + 3

        valve_data = self.build_valve_per_element_data()
        if valve_data:
            for valve_info in valve_data:
                ws.cell(
                    row=current_row, column=1, value=valve_info["valve_name"]
                ).font = _STYLES.title
                current_row += 2
                end_row = self._write_valve_table(ws, valve_info, current_row)
                current_row = end_row + 3

        placeholder_elements = [
            ("Compressors", self.build_compressor_summary_df),
            ("Heat Exchangers", self.build_exchanger_summary_df),
            ("Misc Equipment", self.build_misc_summary_df),
            ("Feeds", self.build_feed_summary_df),
            ("Products", self.build_product_summary_df),
            ("Orifices", self.build_orifice_summary_df),
        ]

        for element_name, builder_func in placeholder_elements:
            df = builder_func()
            if not df.empty:
                ws.cell(
                    row=current_row, column=1, value=f"{element_name} Summary"
                ).font = _STYLES.title
                current_row += 2
                end_row = self._write_placeholder_table(ws, df, current_row, element_name)
                current_row = end_row + 3

        footer_text = (
            "This report is auto-generated from the KORF hydraulic model. "
            "Do not edit this document directly — any changes must be made in the source model "
            f"({source_name}) and the report regenerated."
        )
        footer_cell = ws.cell(row=current_row, column=1, value=footer_text)
        footer_cell.font = _STYLES.footer
        footer_cell.alignment = Alignment(wrap_text=False)
        ws.row_dimensions[current_row].height = 30

    def build_pipe_summary_df(self) -> pd.DataFrame:
        """Build pipe summary DataFrame with governing case based on max velocity.

        Governing case determination:
        - For each pipe, find the case with maximum velocity_in
        - Use that case's data row
        - Add 'Governing Case' column after 'Pipe Name'
        """
        sorted_cases = sorted(self._case_data.keys(), key=lambda c: int(c.number))

        pipe_rows: list[dict] = []
        pipe_names: set[str] = set()

        for case_info in sorted_cases:
            case_data = self._case_data[case_info]
            for pipe in case_data.pipes:
                if pipe.name.startswith("d"):
                    continue
                if pipe.length is not None and pipe.length < 5.0:
                    dp = (
                        abs(pipe.pressure_in - pipe.pressure_out)
                        if pipe.pressure_in is not None and pipe.pressure_out is not None
                        else None
                    )
                    if dp is None or dp <= 0.5:
                        continue
                pipe_names.add(pipe.name)

        for pipe_name in sorted(pipe_names):
            best_case_name = ""
            governing_velcoty: float | None = None
            governing_pipe_data: PipeData | None = None

            for case_info in sorted_cases:
                case_data = self._case_data[case_info]
                pipe = next((p for p in case_data.pipes if p.name == pipe_name), None)
                if pipe is None:
                    continue

                velocity = pipe.velocity_in
                if velocity is None:
                    continue

                if governing_velcoty is None or abs(velocity) > abs(governing_velcoty):
                    governing_velcoty = velocity
                    best_case_name = case_info.name
                    governing_pipe_data = pipe

            if governing_pipe_data is None:
                continue

            row = self._build_pipe_row(governing_pipe_data)
            row["Governing Case"] = best_case_name

            cols = list(row.keys())
            if "Pipe Name" in cols:
                idx = cols.index("Pipe Name")
                cols.insert(idx + 1, "Governing Case")
                cols.remove("Governing Case")
                cols = ["Pipe Name", "Governing Case"] + [
                    c for c in row.keys() if c not in ("Pipe Name", "Governing Case")
                ]
                row = {k: row.get(k) for k in cols}

            pipe_rows.append(row)

        converter = UnitConverter()
        pipe_rows = converter.convert_summary(pipe_rows)

        if not pipe_rows:
            return pd.DataFrame()

        df = pd.DataFrame(pipe_rows)
        cols = list(df.columns)
        if "Pipe Name" in cols and "Governing Case" in cols:
            new_order = ["Pipe Name", "Governing Case"] + [
                c for c in cols if c not in ("Pipe Name", "Governing Case")
            ]
            df = df[new_order]

        return df

    def build_pump_summary_df(self) -> pd.DataFrame:
        """Build pump summary DataFrame with first case as governing.

        Governing case = first case from sorted case list.
        Add 'Governing Case' column after 'Pump Name'.
        """
        sorted_cases = sorted(self._case_data.keys(), key=lambda c: int(c.number))
        if not sorted_cases:
            return pd.DataFrame()

        first_case = sorted_cases[0]
        first_case_name = first_case.name
        case_data = self._case_data[first_case]

        pump_rows: list[dict] = []
        for pump in case_data.pumps:
            row = self._build_pump_row(pump, case_data)
            row["Governing Case"] = first_case_name

            cols = list(row.keys())
            if "Pump Name" in cols:
                cols = ["Pump Name", "Governing Case"] + [
                    c for c in row.keys() if c not in ("Pump Name", "Governing Case")
                ]
                row = {k: row.get(k) for k in cols}

            pump_rows.append(row)

        if not pump_rows:
            return pd.DataFrame()

        df = pd.DataFrame(pump_rows)
        cols = list(df.columns)
        if "Pump Name" in cols and "Governing Case" in cols:
            new_order = ["Pump Name", "Governing Case"] + [
                c for c in cols if c not in ("Pump Name", "Governing Case")
            ]
            df = df[new_order]

        return df

    def _build_pipe_row(self, pd_pipe: PipeData) -> dict:
        """Build a single pipe row dict (mirrors KorfReporter._extract_pipes)."""
        pipe_model = self._find_pipe_in_model(pd_pipe.name)

        criteria_code = ""
        line_number = ""
        rho_v2_min: float | None = None
        rho_v2_max: float | None = None

        if pipe_model and pipe_model.index != 0:
            criteria_code = pipe_model.criteria_code or ""
            from pykorf.app.operation.data_import.line_number import LineNumber

            parsed = LineNumber.parse(pipe_model.notes)
            line_number = parsed.raw_line_number if parsed else ""

            if criteria_code:
                from pykorf.app.operation.integration.sizing_criteria import (
                    code_to_state,
                    lookup_criteria,
                )

                state = code_to_state(criteria_code)
                if state:
                    try:
                        size_inch = (
                            float(pipe_model.diameter_inch) if pipe_model.diameter_inch else None
                        )
                    except (ValueError, TypeError):
                        size_inch = None
                    pressures = pipe_model.pressure or []
                    try:
                        pressure_barg = pressures[0] / 100.0 if pressures else None
                    except (IndexError, TypeError):
                        pressure_barg = None
                    crit = lookup_criteria(state, criteria_code, size_inch, pressure_barg)
                    if crit:
                        rho_v2_min = round(crit.rho_v2_min) if crit.rho_v2_min else None
                        rho_v2_max = round(crit.rho_v2_max) if crit.rho_v2_max is not None else None

        units = pd_pipe.units
        length_unit = units.get("length", "m")
        dp_length_unit = units.get("dp_length", "bar/100m")
        dp_crit_unit = units.get("dp_length_criteria_max", "bar/100m")
        vel_unit = units.get("velocity_in", "m/s")
        vel_max_unit = units.get("velocity_criteria_max", "m/s")
        vel_min_unit = units.get("velocity_criteria_min", "m/s")
        rho_v2_unit = units.get("rho_v2_in", "Pa")

        volume: float | None = None
        if pipe_model and pipe_model.index != 0:
            vol = pipe_model.volume_m3
            volume = round(vol, 4) if vol > 0 else None

        return {
            "Pipe Name": pd_pipe.name,
            "Criteria Code": criteria_code,
            "Line Number": line_number,
            "Line Size": pd_pipe.size or "",
            f"Line Length [{length_unit}]": pd_pipe.length,
            "Volume [m³]": volume,
            f"dP max Criteria [{dp_crit_unit}]": pd_pipe.dp_length_criteria_max,
            f"v min Criteria [{vel_min_unit}]": pd_pipe.velocity_criteria_min,
            f"v max Criteria [{vel_max_unit}]": pd_pipe.velocity_criteria_max,
            "ρV² min Criteria [Pa]": rho_v2_min,  # noqa: RUF001
            "ρV² max Criteria [Pa]": rho_v2_max,  # noqa: RUF001
            f"DP/Length [{dp_length_unit}]": pd_pipe.dp_length,
            f"Velocity [{vel_unit}]": pd_pipe.velocity_in,
            f"ρV² calc [{rho_v2_unit}]": (  # noqa: RUF001
                round(pd_pipe.rho_v2_in) if pd_pipe.rho_v2_in is not None else None
            ),
            "Criteria Check": self._check_pipe_criteria(pd_pipe),
        }

    def _find_pipe_in_model(self, name: str):
        """Find a pipe in the model by name."""
        for idx, pipe in self._model.pipes.items():
            if idx != 0 and pipe.name == name:
                return pipe
        return None

    def _check_pipe_criteria(self, pd_pipe: PipeData) -> str:
        """Check pipe criteria and return PASS/FAIL/JUSTIFIED status."""
        failed = False
        if (
            pd_pipe.dp_length_criteria_max is not None
            and pd_pipe.dp_length is not None
            and pd_pipe.dp_length_criteria_max > 0
        ):
            if abs(pd_pipe.dp_length) > abs(pd_pipe.dp_length_criteria_max):
                failed = True
        if (
            pd_pipe.velocity_criteria_max is not None
            and pd_pipe.velocity_in is not None
            and pd_pipe.velocity_criteria_max > 0
        ):
            if abs(pd_pipe.velocity_in) > abs(pd_pipe.velocity_criteria_max):
                failed = True
        if pd_pipe.velocity_criteria_min is not None and pd_pipe.velocity_in is not None:
            if (
                pd_pipe.velocity_criteria_min > 0
                and abs(pd_pipe.velocity_in) < pd_pipe.velocity_criteria_min
            ):
                failed = True
        if failed:
            if pd_pipe.name in self._justifications:
                return "JUSTIFIED"
            return "FAIL"
        return "PASS"

    def _resolve_pipe_columns(self, df: pd.DataFrame, keys: list[str]) -> list[str]:
        """Resolve partial column keys to actual DataFrame column names.

        Does exact match first, then prefix match (e.g. "Velocity" → "Velocity [m/s]").
        """
        available = set(df.columns)
        resolved: list[str] = []
        for key in keys:
            if key in available:
                resolved.append(key)
            else:
                for col in df.columns:
                    if col.startswith(key + " [") or col == key:
                        resolved.append(col)
                        break
        return resolved

    def _filter_pipe_columns(self, df: pd.DataFrame, pipe_columns: list[str]) -> pd.DataFrame:
        """Filter pipe DataFrame to only include requested columns.

        Always includes 'Pipe Name' and 'Criteria Check' regardless of selection.
        Preserves the original DataFrame column order.
        """
        always_keys = {"Pipe Name", "Governing Case", "Criteria Check"}
        resolved = self._resolve_pipe_columns(df, pipe_columns)
        always_resolved = self._resolve_pipe_columns(df, list(always_keys))
        selected_set = set(resolved) | set(always_resolved)
        ordered = [c for c in df.columns if c in selected_set]
        if not ordered:
            return df
        return df[ordered]

    def _build_pump_row(self, pump: PumpData, case_data: KorfCaseData) -> dict:
        """Build a single pump row dict (mirrors KorfReporter._extract_pumps)."""
        inlet_pipe = next((p for p in case_data.pipes if p.name == pump.pipe_inlet), None)
        temperature = inlet_pipe.temperature_out if inlet_pipe else None
        viscosity = inlet_pipe.viscosity if inlet_pipe else None

        return {
            "Pump Name": pump.name,
            "Section_Liquid Characteristics": "Liquid Characteristics",
            "Vapour Pressure [bara]": pump.vapour_pressure,
            "Density [kg/m³]": pump.density,
            "Viscosity [cP]": viscosity,
            "Section_Operating Conditions": "Operating Conditions",
            "Pump Datum Elevation [m]": pump.elevation,
            "Pumping Temperature [°C]": temperature,
            "Volumetric Flow [m³/h]": pump.vol_flow,
            "Discharge Pressure [barg]": pump.pressure_out,
            "Suction Pressure [barg]": pump.pressure_in,
            "Differential Pressure [bar]": pump.dp,
            "Differential Head [m]": pump.head,
            "NPSH Available [m]": pump.npsha_calc,
            "Shaft Power [kW]": pump.power,
            "Efficiency [%]": round(pump.efficiency * 100, 1)
            if pump.efficiency is not None
            else None,
            "Hydraulic Power [kW]": pump.hydraulic_power,
            "Section_Performance Characteristics": "Performance Characteristics",
            "Raise to Shutoff DP []": pump.raise_to_shutoff_dp,
            "Suc Vessel Design Pressure [barg]": pump.vessel_pressure,
            "Suc Vessel Max Level [m]": pump.vessel_max_level,
            "Shut-Off DP [bar]": pump.shutoff_dp,
            "Suction Max Pressure [barg]": pump.suction_max_pressure,
            "Discharge Shut-Off Pressure [barg]": pump.shutoff_pressure,
        }

    def build_valve_per_element_data(self) -> list[dict]:
        """Build per-valve data for individual tables.

        Returns a list of dicts, each containing:
        - valve_name: str
        - parameters: list[str]
        - units: list[str]
        - case_values: dict[case_name, list[value]]
        """
        sorted_cases = sorted(self._case_data.keys(), key=lambda c: int(c.number))
        case_names = [c.name for c in sorted_cases]

        valve_names: set[str] = set()
        for case_info in sorted_cases:
            case_data = self._case_data[case_info]
            for valve in case_data.valves:
                valve_names.add(valve.name)

        result: list[dict] = []
        for valve_name in sorted(valve_names):
            valve_info = {
                "valve_name": valve_name,
                "parameters": [p[0] for p in _VALVE_PARAMETERS],
                "units": [p[1] for p in _VALVE_PARAMETERS],
                "case_values": {},
            }

            for case_info in sorted_cases:
                case_data = self._case_data[case_info]
                valve = next((v for v in case_data.valves if v.name == valve_name), None)
                if valve is None:
                    valve_info["case_values"][case_info.name] = [None] * len(_VALVE_PARAMETERS)
                    continue

                inlet_pipe = next((p for p in case_data.pipes if p.name == valve.pipe_inlet), None)
                flow_rate = inlet_pipe.mass_flow if inlet_pipe else None

                values = [
                    flow_rate,
                    valve.pressure_in,
                    valve.dp,
                    None,
                ]
                valve_info["case_values"][case_info.name] = values

            result.append(valve_info)

        return result

    def _write_pipe_table(
        self, ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int = 1
    ) -> int:
        """Write pipe summary table to worksheet."""
        descriptions, units = parse_headers(df.columns)

        write_two_level_headers(ws, start_row, start_col, descriptions, units)

        data_start_row = start_row + 2
        criteria_col_idx = None
        for i, col in enumerate(df.columns):
            if col == "Criteria Check":
                criteria_col_idx = start_col + i
                break

        for r_idx, row_data in enumerate(df.values, start=data_start_row):
            for c_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = _STYLES.data
                if c_idx != start_col:
                    cell.alignment = _STYLES.data_center_align
                if c_idx > start_col:
                    col_name = df.columns[c_idx - start_col]
                    apply_number_format(cell, col_name)
                if criteria_col_idx is not None and c_idx == criteria_col_idx:
                    if val == "FAIL":
                        apply_fail_format(cell)
                        cell.fill = _STYLES.fail_fill
                    elif val == "JUSTIFIED":
                        apply_justified_format(cell)

        last_row = data_start_row + len(df) - 1
        apply_table_borders(ws, start_row, last_row, len(descriptions), start_col)

        return last_row

    def _write_pump_table(
        self, ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int = 1
    ) -> int:
        """Write pump summary table (transposed format with section markers)."""
        descriptions = list(df.columns)
        pump_names = df.iloc[:, 0].tolist()

        num_cols = 2 + len(pump_names)

        write_transposed_header(ws, start_row, start_col, ["Parameter", "Unit", *pump_names])

        current_row = start_row + 1
        units_row_idx = None
        for i, col in enumerate(descriptions):
            if "Unit" in col:
                units_row_idx = i
                break

        for p_idx in range(1, len(descriptions)):
            if p_idx == units_row_idx:
                continue

            param_name = descriptions[p_idx]
            row_values = df.iloc[:, p_idx].tolist()

            is_section_marker = (
                param_name
                and param_name.startswith("Section_")
                and len(row_values) > 0
                and row_values[0] == param_name[8:]
            )

            if is_section_marker:
                section_name = param_name[8:]
                write_section_marker(ws, current_row, start_col, num_cols, section_name)
            else:
                unit_match = parse_header_unit(param_name)
                param_display = unit_match[0] if unit_match else param_name
                unit_display = f"[{unit_match[1]}]" if unit_match else ""

                cell_p = ws.cell(row=current_row, column=start_col, value=param_display)
                cell_p.font = _STYLES.header
                cell_p.fill = _STYLES.header_fill

                cell_u = ws.cell(row=current_row, column=start_col + 1, value=unit_display)
                cell_u.font = _STYLES.unit

                for v_idx, val in enumerate(row_values, start=start_col + 2):
                    cell_v = ws.cell(row=current_row, column=v_idx, value=val)
                    cell_v.font = _STYLES.data
                    cell_v.alignment = _STYLES.data_center_align
                    apply_number_format(cell_v, param_display)

            current_row += 1

        last_row = current_row - 1
        apply_table_borders(ws, start_row, last_row, num_cols, start_col)

        return last_row

    def _write_valve_table(
        self, ws: Worksheet, valve_info: dict, start_row: int, start_col: int = 1
    ) -> int:
        """Write a single valve's per-case table."""
        case_names = list(valve_info["case_values"].keys())
        num_cols = 2 + len(case_names)

        ws.row_dimensions[start_row].height = 30
        headers = ["Parameter", "Unit", *case_names]
        for c_idx, val in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=c_idx, value=val)
            cell.font = _STYLES.header
            cell.fill = _STYLES.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row = start_row + 1
        for i, (param, unit) in enumerate(_VALVE_PARAMETERS):
            cell_p = ws.cell(row=current_row, column=start_col, value=param)
            cell_p.font = _STYLES.header
            cell_p.fill = _STYLES.header_fill

            cell_u = ws.cell(row=current_row, column=start_col + 1, value=f"[{unit}]")
            cell_u.font = _STYLES.unit

            for case_idx, case_name in enumerate(case_names):
                val = valve_info["case_values"][case_name][i]
                cell_v = ws.cell(row=current_row, column=start_col + 2 + case_idx, value=val)
                cell_v.font = _STYLES.data
                cell_v.alignment = Alignment(horizontal="center")
                apply_number_format(cell_v, param)

            current_row += 1

        last_row = current_row - 1
        apply_table_borders(ws, start_row, last_row, num_cols, start_col)

        return last_row

    def _write_placeholder_table(
        self, ws: Worksheet, df: pd.DataFrame, start_row: int, element_name: str, start_col: int = 1
    ) -> int:
        """Write placeholder table for elements yet to be implemented."""
        if df.empty:
            ws.cell(
                row=start_row,
                column=start_col,
                value=f"{element_name}: TODO - Implementation pending",
            )
            return start_row

        descriptions, units = parse_headers(df.columns)

        ws.row_dimensions[start_row].height = 30
        for c_idx, (desc, unit) in enumerate(
            zip(descriptions, units, strict=True), start=start_col
        ):
            cell_desc = ws.cell(row=start_row, column=c_idx, value=desc)
            cell_desc.font = _STYLES.header
            cell_desc.fill = _STYLES.header_fill
            cell_desc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            cell_unit = ws.cell(row=start_row + 1, column=c_idx, value=unit)
            cell_unit.font = _STYLES.unit
            cell_unit.fill = _STYLES.header_fill
            cell_unit.alignment = Alignment(horizontal="center")

        data_start_row = start_row + 2
        for r_idx, row_data in enumerate(df.values, start=data_start_row):
            for c_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = _STYLES.data
                if c_idx != start_col:
                    cell.alignment = Alignment(horizontal="center")
                if c_idx > start_col:
                    apply_number_format(cell, df.columns[c_idx - start_col])

        last_row = data_start_row + len(df) - 1
        apply_table_borders(ws, start_row, last_row, len(descriptions), start_col)

        return last_row

    def build_compressor_summary_df(self) -> pd.DataFrame:
        """TODO: Build compressor summary DataFrame."""
        return pd.DataFrame()

    def build_exchanger_summary_df(self) -> pd.DataFrame:
        """TODO: Build heat exchanger summary DataFrame."""
        return pd.DataFrame()

    def build_misc_summary_df(self) -> pd.DataFrame:
        """TODO: Build misc equipment summary DataFrame."""
        return pd.DataFrame()

    def build_feed_summary_df(self) -> pd.DataFrame:
        """TODO: Build feed summary DataFrame."""
        return pd.DataFrame()

    def build_product_summary_df(self) -> pd.DataFrame:
        """TODO: Build product summary DataFrame."""
        return pd.DataFrame()

    def build_orifice_summary_df(self) -> pd.DataFrame:
        """TODO: Build orifice summary DataFrame."""
        return pd.DataFrame()
