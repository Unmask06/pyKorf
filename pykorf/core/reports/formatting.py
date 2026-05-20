"""Centralized Excel formatting for all report generators.

Defines font styles, fills, borders, number format rules, and shared
helper functions used by ResultExporter, MultiCaseSummaryBuilder, and
any future report writers.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ReportStyles:
    """Central registry of all report cell styles.

    Usage::

        styles = ReportStyles()
        cell.font = styles.model_title
        cell.fill = styles.header_fill
    """

    model_title = Font(bold=True, size=18, color="003366")
    title = Font(bold=True, size=14, color="003366")
    header = Font(bold=True, size=10)
    unit = Font(italic=True, color="555555", size=10)
    data = Font(size=10)
    footer = Font(italic=True, size=9, color="888888")

    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_side = Side(style="thin")
    thick_side = Side(style="medium")

    fail_font = Font(bold=True, size=10, color="9C0006")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    section_font = Font(bold=True, italic=True, size=11, color="003366")
    section_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left", vertical="center")


# Module-level singleton for convenience
_STYLES = ReportStyles()


# ---- Header Parsing ----------------------------------------------------------

_HEADER_PATTERN = re.compile(r"^(.*?) \[(.*?)\]$")


def parse_headers(columns: list[str]) -> tuple[list[str], list[str]]:
    """Split column names into (description, unit) tuples.

    >>> parse_headers(["Pipe Name", "Line Length [m]", "DP/Length [bar/100m]"])
    (["Pipe Name", "Line Length", "DP/Length"], ["", "[m]", "[bar/100m]"])
    """
    descriptions, units = [], []
    for col in columns:
        match = _HEADER_PATTERN.match(col)
        if match:
            descriptions.append(match.group(1))
            units.append(f"[{match.group(2)}]")
        else:
            descriptions.append(col)
            units.append("")
    return descriptions, units


def parse_header_unit(col_name: str) -> tuple[str, str] | None:
    """Parse a single column name into (description, unit) if it has ``[unit]`` format."""
    match = _HEADER_PATTERN.match(col_name)
    if match:
        return match.group(1), match.group(2)
    return None


# ---- Number Formatting -------------------------------------------------------

# Map of column name keywords to Excel number_format.
# Each entry is (tuple_of_keywords, format).
# Rules are applied in order; LAST match wins.
# Generic rules first, specific overrides after.
NUMBER_FORMAT_RULES: list[tuple[tuple[str, ...], str]] = [
    # ── Generic rules (applied to all matching columns) ──────────────
    (("\u03c1V\u00b2",), "#,##0"),
    (("Differential Head", "Temperature", "Flowrate"), "#,##0"),
    (("Velocity", "Pressure", "Differential Pressure"), "#0.00"),
    (("dP",), "#0.00"),
    (("Raise", "Shut-Off", "NPSH", "Power", "Volumetric Flow"), "#0.0"),
]


def apply_number_format(cell: Any, header_name: str) -> None:
    """Apply the registered number format to *cell* if *header_name* matches any keyword.

    Matching is case-insensitive. LAST matching rule wins so that
    specific overrides placed below supersede generic rules.
    """
    lower = header_name.lower()
    matched_fmt: str | None = None
    for keywords, fmt in NUMBER_FORMAT_RULES:
        if any(kw.lower() in lower for kw in keywords):
            matched_fmt = fmt
    if matched_fmt is not None:
        cell.number_format = matched_fmt


# ---- Column Widths -----------------------------------------------------------

_COL_A_WIDTH = 25
_OTHER_COL_WIDTH = 15


def apply_column_widths(ws: Any, num_cols: int, start_col: int = 1) -> None:
    """Set fixed column widths: col A = 25, others = 15."""
    ws.column_dimensions[get_column_letter(start_col)].width = _COL_A_WIDTH
    for c in range(start_col + 1, start_col + num_cols):
        ws.column_dimensions[get_column_letter(c)].width = _OTHER_COL_WIDTH


# ---- Borders -----------------------------------------------------------------


def apply_table_borders(
    ws: Any, start_row: int, end_row: int, num_cols: int, start_col: int = 1
) -> None:
    """Apply thin internal borders and a thick outer border to a table block."""
    thin_s = _STYLES.thin_side
    thick_s = _STYLES.thick_side

    end_col = start_col + num_cols - 1
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            left = thick_s if c == start_col else thin_s
            right = thick_s if c == end_col else thin_s
            top = thick_s if r == start_row else thin_s
            bottom = thick_s if r == end_row else thin_s
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


# ---- FAIL Cell Formatting ----------------------------------------------------


def apply_fail_format(cell: Any) -> None:
    """Apply red bold font + red fill to a FAIL criteria cell."""
    cell.font = _STYLES.fail_font
    cell.fill = _STYLES.fail_fill


# ---- Section Marker (Transposed Tables) --------------------------------------


def write_section_marker(
    ws: Any, row: int, start_col: int, num_cols: int, section_name: str
) -> None:
    """Write a styled section separator row spanning all columns."""
    end_col = start_col + num_cols - 1
    cell = ws.cell(row=row, column=start_col, value=section_name)
    cell.font = _STYLES.section_font
    cell.fill = _STYLES.section_fill
    cell.alignment = _STYLES.left_align
    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )
    ws.row_dimensions[row].height = 25


# ---- Two-Level Header Writers ------------------------------------------------


def write_two_level_headers(
    ws: Any,
    row: int,
    start_col: int,
    descriptions: list[str],
    units: list[str],
) -> None:
    """Write a two-row header (row 1 = description, row 2 = unit)."""
    ws.row_dimensions[row].height = 30
    for c_idx, (desc, unit) in enumerate(zip(descriptions, units, strict=True), start=start_col):
        cell_desc = ws.cell(row=row, column=c_idx, value=desc)
        cell_desc.font = _STYLES.header
        cell_desc.fill = _STYLES.header_fill
        cell_desc.alignment = _STYLES.header_align

        cell_unit = ws.cell(row=row + 1, column=c_idx, value=unit)
        cell_unit.font = _STYLES.unit
        cell_unit.fill = _STYLES.header_fill
        cell_unit.alignment = Alignment(horizontal="center")


def write_transposed_header(ws: Any, row: int, start_col: int, header_names: list[str]) -> None:
    """Write a single-row transposed table header."""
    ws.row_dimensions[row].height = 30
    for c_idx, val in enumerate(header_names, start=start_col):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = _STYLES.header
        cell.fill = _STYLES.header_fill
        cell.alignment = _STYLES.header_align
