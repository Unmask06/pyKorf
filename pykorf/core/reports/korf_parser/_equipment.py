"""Equipment sheet section parsers for KORF Excel reports."""

from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from pykorf.core.reports.korf_parser._columns import (
    _COMPRESSOR_COL_MAP,
    _EXCHANGER_COL_MAP,
    _FEED_COL_MAP,
    _MISC_COL_MAP,
    _NPSH_COL_MAP,
    _ORIFICE_COL_MAP,
    _PRODUCT_COL_MAP,
    _PUMP_COL_MAP,
    _SECTION_MARKERS_EQUIPMENT,
    _SHUTOFF_COL_MAP,
    _VALVE_COL_MAP,
)
from pykorf.core.reports.korf_parser._utils import _safe_float, _safe_str
from pykorf.core.reports.korf_parser.models import (
    CompressorData,
    ExchangerData,
    FeedData,
    MiscEquipmentData,
    OrificeData,
    ProductData,
    PumpData,
    ValveData,
)


def _extract_row_data_direct(
    ws: Worksheet,
    data_row: int,
    col_map: dict[str, int],
) -> dict[str, Any]:
    """Extract data from a single row using explicit column indices."""
    result: dict[str, Any] = {}
    for param, col in col_map.items():
        result[param] = ws.cell(row=data_row, column=col).value
    return result


# ── EQUIPMENT SHEET PARSER ───────────────────────────────────────────


def _parse_equipment_sheet(ws: Worksheet) -> dict[str, list]:
    """Parse an Equipment sheet to extract all element data."""
    result: dict[str, list] = {
        "feeds": [],
        "products": [],
        "orifices": [],
        "valves": [],
        "pumps": [],
        "compressors": [],
        "exchangers": [],
        "misc_equipment": [],
    }

    row = 1
    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)

        if cell_a == "FEEDS":
            feeds, row = _parse_feeds_section(ws, row)
            result["feeds"] = feeds
            continue
        elif cell_a == "PRODUCTS":
            products, row = _parse_products_section(ws, row)
            result["products"] = products
            continue
        elif cell_a == "ORIFICES":
            orifices, row = _parse_orifices_section(ws, row)
            result["orifices"] = orifices
            continue
        elif cell_a == "VALVES":
            valves, row = _parse_valves_section(ws, row)
            result["valves"] = valves
            continue
        elif cell_a == "PUMPS":
            pumps, row = _parse_pumps_section(ws, row)
            result["pumps"] = pumps
            continue
        elif cell_a == "COMPRESSORS":
            compressors, row = _parse_compressors_section(ws, row)
            result["compressors"] = compressors
            continue
        elif cell_a == "EXCHANGERS":
            exchangers, row = _parse_exchangers_section(ws, row)
            result["exchangers"] = exchangers
            continue
        elif cell_a == "MISCELLANEOUS":
            misc, row = _parse_misc_section(ws, row)
            result["misc_equipment"] = misc
            continue

        row += 1

    return result


def _parse_feeds_section(ws: Worksheet, start_row: int) -> tuple[list[FeedData], int]:
    """Parse FEEDS section from Equipment sheet using explicit column indices."""
    feeds: list[FeedData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _FEED_COL_MAP)
        feeds.append(
            FeedData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                elevation=_safe_float(data.get("elevation")),
                fluid_level=_safe_float(data.get("fluid_level")),
                pressure=_safe_float(data.get("pressure")),
            )
        )
        row += 1

    return feeds, row


def _parse_products_section(ws: Worksheet, start_row: int) -> tuple[list[ProductData], int]:
    """Parse PRODUCTS section from Equipment sheet using explicit column indices."""
    products: list[ProductData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _PRODUCT_COL_MAP)
        products.append(
            ProductData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                elevation=_safe_float(data.get("elevation")),
                fluid_level=_safe_float(data.get("fluid_level")),
                pressure=_safe_float(data.get("pressure")),
            )
        )
        row += 1

    return products, row


def _parse_orifices_section(ws: Worksheet, start_row: int) -> tuple[list[OrificeData], int]:
    """Parse ORIFICES section from Equipment sheet using explicit column indices."""
    orifices: list[OrificeData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _ORIFICE_COL_MAP)
        orifices.append(
            OrificeData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                type=_safe_str(data.get("type")),
                bore=_safe_float(data.get("bore")),
                beta=_safe_float(data.get("beta")),
                dp_flange_tap=_safe_float(data.get("dp_flange")),
                dp_pipe_tap=_safe_float(data.get("dp_pipe")),
                pressure_in=_safe_float(data.get("pressure_in")),
                pressure_out=_safe_float(data.get("pressure_out")),
                pipe_inlet=_safe_str(data.get("pipe_inlet")),
                pipe_outlet=_safe_str(data.get("pipe_outlet")),
            )
        )
        row += 1

    return orifices, row


def _parse_valves_section(ws: Worksheet, start_row: int) -> tuple[list[ValveData], int]:
    """Parse VALVES section from Equipment sheet using explicit column indices."""
    valves: list[ValveData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _VALVE_COL_MAP)
        valves.append(
            ValveData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                type=_safe_str(data.get("type")),
                cv=_safe_float(data.get("cv")),
                dp=_safe_float(data.get("dp")),
                pressure_in=_safe_float(data.get("pressure_in")),
                pressure_out=_safe_float(data.get("pressure_out")),
                pipe_inlet=_safe_str(data.get("pipe_inlet")),
                pipe_outlet=_safe_str(data.get("pipe_outlet")),
            )
        )
        row += 1

    return valves, row


def _parse_pumps_section(ws: Worksheet, start_row: int) -> tuple[list[PumpData], int]:
    """Parse PUMPS section including NPSH and SHUT OFF PRESSURE subsections."""
    pumps: list[PumpData] = []
    unit_row = start_row + 2
    data_row = start_row + 4

    pump_units: dict[str, str] = {}
    if unit_row <= ws.max_row:
        for field_name, col_idx in _PUMP_COL_MAP.items():
            cell_val = ws.cell(row=unit_row, column=col_idx).value
            if cell_val:
                pump_units[field_name] = _safe_str(cell_val)

    while data_row <= ws.max_row:
        cell_a = ws.cell(row=data_row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if (
            cell_a_str in ("NPSH", "SHUT OFF PRESSURE", "CURVES")
            or cell_a_str in _SECTION_MARKERS_EQUIPMENT
        ):
            break
        if not cell_a_str:
            data_row += 1
            continue

        data = _extract_row_data_direct(ws, data_row, _PUMP_COL_MAP)
        pumps.append(
            PumpData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                elevation=_safe_float(data.get("elevation")),
                efficiency=_safe_float(data.get("efficiency")),
                power=_safe_float(data.get("power")),
                flow=_safe_float(data.get("flow")),
                density=_safe_float(data.get("density")),
                vol_flow=_safe_float(data.get("vol_flow")),
                head=_safe_float(data.get("head")),
                dp=_safe_float(data.get("dp")),
                pressure_in=_safe_float(data.get("pressure_in")),
                pressure_out=_safe_float(data.get("pressure_out")),
                pipe_inlet=_safe_str(data.get("pipe_inlet")),
                pipe_outlet=_safe_str(data.get("pipe_outlet")),
                pressure_in_unit=pump_units.get("pressure_in", ""),
                density_unit=pump_units.get("density", ""),
            )
        )
        data_row += 1

    # Parse NPSH subsection
    while data_row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=data_row, column=1).value)
        if cell_a == "NPSH":
            npsh_unit_row = data_row + 2
            npsh_data_row = data_row + 4
            npsh_units: dict[str, str] = {}
            if npsh_unit_row <= ws.max_row:
                for field_name, col_idx in _NPSH_COL_MAP.items():
                    cell_val = ws.cell(row=npsh_unit_row, column=col_idx).value
                    if cell_val:
                        npsh_units[field_name] = _safe_str(cell_val)
            if npsh_data_row <= ws.max_row:
                for pump in pumps:
                    data = _extract_row_data_direct(ws, npsh_data_row, _NPSH_COL_MAP)
                    if _safe_str(data.get("name")) == pump.name:
                        pump.npsha = _safe_float(data.get("npsha"))
                        pump.npshr = _safe_float(data.get("npshr"))
                        pump.vapour_pressure = _safe_float(data.get("vapour_pressure"))
                        pump.vapour_pressure_unit = npsh_units.get("vapour_pressure", "")
                        pump.contigency = _safe_float(data.get("contigency")) or 0.0
            data_row += 1
            continue
        if cell_a in ("SHUT OFF PRESSURE", "CURVES") or cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        data_row += 1

    # Parse SHUT OFF PRESSURE subsection
    while data_row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=data_row, column=1).value)
        if cell_a == "SHUT OFF PRESSURE":
            shutoff_data_row = data_row + 4
            if shutoff_data_row <= ws.max_row:
                for pump in pumps:
                    data = _extract_row_data_direct(ws, shutoff_data_row, _SHUTOFF_COL_MAP)
                    if _safe_str(data.get("name")) == pump.name:
                        pump.vessel_pressure = _safe_float(data.get("vessel_pressure"))
                        pump.vessel_max_level = _safe_float(data.get("vessel_max_level"))
                        pump.suction_max_pressure = _safe_float(data.get("suction_max_pressure"))
                        pump.shutoff_dp = _safe_float(data.get("shutoff_dp"))
                        pump.shutoff_pressure = _safe_float(data.get("shutoff_pressure"))
                        pump.raise_to_shutoff_dp = _safe_float(data.get("raise_to_shutoff_dp"))
            data_row += 1
            continue
        if cell_a in ("CURVES",) or cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        data_row += 1

    return pumps, data_row


def _parse_compressors_section(ws: Worksheet, start_row: int) -> tuple[list[CompressorData], int]:
    """Parse COMPRESSORS section from Equipment sheet using explicit column indices."""
    compressors: list[CompressorData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _COMPRESSOR_COL_MAP)
        compressors.append(
            CompressorData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                type=_safe_str(data.get("type")),
                dp=_safe_float(data.get("dp")),
                pressure_in=_safe_float(data.get("pressure_in")),
                pressure_out=_safe_float(data.get("pressure_out")),
                flow=_safe_float(data.get("flow")),
                power=_safe_float(data.get("power")),
                efficiency=_safe_float(data.get("efficiency")),
                head=_safe_float(data.get("head")),
            )
        )
        row += 1

    return compressors, row


def _parse_exchangers_section(ws: Worksheet, start_row: int) -> tuple[list[ExchangerData], int]:
    """Parse EXCHANGERS section from Equipment sheet using explicit column indices."""
    exchangers: list[ExchangerData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _EXCHANGER_COL_MAP)
        exchangers.append(
            ExchangerData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                type=_safe_str(data.get("type")),
                side=_safe_str(data.get("side")),
                elevation_in=_safe_float(data.get("elevation_in")),
                duty=_safe_float(data.get("duty")),
                dp=_safe_float(data.get("dp")),
                pressure_in=_safe_float(data.get("pressure_in")),
                pressure_out=_safe_float(data.get("pressure_out")),
                pipe_inlet=_safe_str(data.get("pipe_inlet")),
                pipe_outlet=_safe_str(data.get("pipe_outlet")),
            )
        )
        row += 1

    return exchangers, row


def _parse_misc_section(ws: Worksheet, start_row: int) -> tuple[list[MiscEquipmentData], int]:
    """Parse MISCELLANEOUS section from Equipment sheet using explicit column indices."""
    misc: list[MiscEquipmentData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = ws.cell(row=row, column=1).value
        cell_a_str = _safe_str(cell_a)
        if cell_a_str in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a_str:
            row += 1
            continue

        data = _extract_row_data_direct(ws, row, _MISC_COL_MAP)
        misc.append(
            MiscEquipmentData(
                name=_safe_str(data.get("name")),
                description=_safe_str(data.get("description")),
                elevation_in=_safe_float(data.get("elevation_in")),
                dp=_safe_float(data.get("dp")),
                pressure_in=_safe_float(data.get("pressure_in")),
                pressure_out=_safe_float(data.get("pressure_out")),
                pipe_inlet=_safe_str(data.get("pipe_inlet")),
                pipe_outlet=_safe_str(data.get("pipe_outlet")),
            )
        )
        row += 1

    return misc, row
