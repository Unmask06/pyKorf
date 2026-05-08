"""Parsing logic for KORF Excel reports."""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pykorf.core.reports.korf_parser.models import (
    CaseInfo,
    KorfCaseData,
    PipeData,
    ValidationEntry,
)

from pykorf.core.reports.korf_parser._equipment import _parse_equipment_sheet
from pykorf.core.reports.korf_parser._utils import _safe_float, _safe_str

_logger = logging.getLogger(__name__)

_SHEET_PATTERN = re.compile(r"^(Title|Profile|Piping|Equipment)-(\d+)\s+(.+)$")


def _parse_cases(ws_names: list[str]) -> dict[tuple[str, str], dict[str, str]]:
    """Parse sheet names to discover cases."""
    cases: dict[tuple[str, str], dict[str, str]] = {}
    for name in ws_names:
        m = _SHEET_PATTERN.match(name)
        if m:
            sheet_type, case_num, case_name = m.groups()
            key = (case_num, case_name)
            cases.setdefault(key, {})[sheet_type] = name
    return cases


# ── TITLE SHEET PARSER ──────────────────────────────────────────────


def _parse_title_sheet(ws: Worksheet) -> tuple[list[ValidationEntry], str]:
    """Parse Title sheet to extract validation issues and run message."""
    validations: list[ValidationEntry] = []
    run_message = ""
    in_warnings = False

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value

        if cell_a and "WARNINGS AND ERRORS" in str(cell_a).upper():
            in_warnings = True
            continue

        if in_warnings and cell_a:
            msg = str(cell_a).strip()
            if msg:
                validations.append(ValidationEntry(message=msg))
            continue

        if cell_a and str(cell_a).strip() == "Run Message":
            run_message = str(cell_b).strip() if cell_b else ""

    return validations, run_message


# ── PIPING SHEET PARSER ──────────────────────────────────────────────

_PIPE_ROW_MAP: dict[tuple[str, str, str], str] = {
    ("Total", "Flow", ""): "mass_flow",
    ("", "Pressure", "In"): "pressure_in",
    ("", "Pressure", "Out"): "pressure_out",
    ("", "Temperature", "In"): "temperature_in",
    ("", "Temperature", "Out"): "temperature_out",
    ("", "Density (No slip)", "In"): "density_in",
    ("", "Density (No slip)", "Out"): "density_out",
    ("", "Volumetric Flow", "Avg"): "vol_flow",
    ("Size", "", ""): "size",
    ("Length", "", ""): "length",
    ("Schedule", "", ""): "schedule",
    ("Overall", "", ""): "dp_overall",
    ("Friction", "", ""): "dp_friction",
    ("Elevation", "", ""): "dp_elevation",
    ("dP/Length", "", ""): "dp_length",
    ("Velocity", "", "In"): "velocity_in",
    ("dP/Length", "Target", "Max"): "dp_length_criteria_max",
    ("Velocity", "Target", "Max"): "velocity_criteria_max",
}

_VISCOSITY_SPELLINGS = {"Viscocity (No slip)", "Viscosity (No slip)"}
_RHO_V2_SPELLINGS = {"Rho.V^2", "Rho.V²", "ρV²"}  # noqa: RUF001

_UNIT_STANDARDIZE: dict[str, str] = {
    "kg/m3.(m/s": "Pa",
    "kg/m3.(m/s]": "Pa",
    "kg/m3.(m/s)^2": "Pa",
    "kg/m3(m/s)^2": "Pa",
    "barg": "barg",
    "bar": "bar",
    "bar/100m": "bar/100m",
    "m": "m",
    "m/s": "m/s",
    "kg/h": "kg/h",
    "m3/h": "m³/h",
    "cP": "cP",
    "C": "°C",
    "kg/m3": "kg/m³",
}


def _standardize_unit(raw_unit: str) -> str:
    """Standardize KORF unit strings to clean format."""
    if not raw_unit:
        return ""
    unit = raw_unit.strip()
    if unit in _UNIT_STANDARDIZE:
        return _UNIT_STANDARDIZE[unit]
    return unit


_REPORT_FIELDS: set[str] = {
    "length",
    "dp_length",
    "dp_length_criteria_max",
    "velocity_in",
    "velocity_criteria_max",
    "velocity_criteria_min",
    "rho_v2_in",
    "dp_overall",
    "dp_friction",
    "dp_elevation",
    "pressure_in",
    "pressure_out",
    "temperature_in",
    "temperature_out",
    "density_in",
    "density_out",
    "viscosity",
    "mass_flow",
    "vol_flow",
}

_CONTINUATION_FIELDS: dict[str, str] = {
    "Pressure": "pressure_out",
    "Temperature": "temperature_out",
    "Density (No slip)": "density_out",
}


def _parse_piping_sheet(ws: Worksheet) -> list[PipeData]:
    """Parse a Piping sheet to extract pipe result data using optimized lookup dict."""
    pipe_names: list[str] = []
    pipe_data: dict[str, dict[str, Any]] = {}
    pipe_units: dict[str, dict[str, str]] = {}

    for col in range(5, ws.max_column + 1):
        name = _safe_str(ws.cell(row=2, column=col).value)
        if name:
            pipe_names.append(name)
            pipe_data[name] = {}
            pipe_units[name] = {}

    if not pipe_names:
        return []

    prev_row_key: tuple[str, str, str] = ("", "", "")
    prev_field_name: str = ""

    for row_idx in range(5, ws.max_row + 1):
        col_a = _safe_str(ws.cell(row=row_idx, column=1).value)
        col_b = _safe_str(ws.cell(row=row_idx, column=2).value)
        col_c = _safe_str(ws.cell(row=row_idx, column=3).value)
        col_d_raw = _safe_str(ws.cell(row=row_idx, column=4).value)
        col_d = _standardize_unit(col_d_raw)

        key = (col_a, col_b, col_c)

        if col_b in _VISCOSITY_SPELLINGS:
            for i, pname in enumerate(pipe_names):
                val = _safe_float(ws.cell(row=row_idx, column=5 + i).value)
                pipe_data[pname]["viscosity"] = val
                if "viscosity" in _REPORT_FIELDS:
                    pipe_units[pname]["viscosity"] = col_d
            prev_row_key = key
            prev_field_name = "viscosity"
            continue

        if col_a in _RHO_V2_SPELLINGS and col_c == "In":
            for i, pname in enumerate(pipe_names):
                val = _safe_float(ws.cell(row=row_idx, column=5 + i).value)
                pipe_data[pname]["rho_v2_in"] = val
                if "rho_v2_in" in _REPORT_FIELDS:
                    pipe_units[pname]["rho_v2_in"] = col_d
            prev_row_key = key
            prev_field_name = "rho_v2_in"
            continue

        if key in _PIPE_ROW_MAP:
            field_name = _PIPE_ROW_MAP[key]
            for i, pname in enumerate(pipe_names):
                if field_name in ("size", "schedule"):
                    val = _safe_str(ws.cell(row=row_idx, column=5 + i).value)
                else:
                    val = _safe_float(ws.cell(row=row_idx, column=5 + i).value)
                pipe_data[pname][field_name] = val
                if field_name in _REPORT_FIELDS and field_name not in ("size", "schedule"):
                    pipe_units[pname][field_name] = col_d
            prev_row_key = key
            prev_field_name = col_b if col_b else col_a
            continue

        if key == ("", "", "Min") and prev_row_key == ("Velocity", "Target", "Max"):
            for i, pname in enumerate(pipe_names):
                val = _safe_float(ws.cell(row=row_idx, column=5 + i).value)
                pipe_data[pname]["velocity_criteria_min"] = val
                if "velocity_criteria_min" in _REPORT_FIELDS:
                    pipe_units[pname]["velocity_criteria_min"] = col_d
            continue

        if key == ("", "", "Out") and prev_field_name in _CONTINUATION_FIELDS:
            field_name = _CONTINUATION_FIELDS[prev_field_name]
            for i, pname in enumerate(pipe_names):
                val = _safe_float(ws.cell(row=row_idx, column=5 + i).value)
                pipe_data[pname][field_name] = val
                if field_name in _REPORT_FIELDS:
                    pipe_units[pname][field_name] = col_d
            continue

        prev_row_key = key
        prev_field_name = col_b if col_b else col_a

    results: list[PipeData] = []
    for pname in pipe_names:
        d = pipe_data.get(pname, {})
        u = pipe_units.get(pname, {})
        results.append(
            PipeData(
                name=pname,
                length=d.get("length"),
                size=d.get("size"),
                schedule=d.get("schedule"),
                dp_length=d.get("dp_length"),
                velocity_in=d.get("velocity_in"),
                rho_v2_in=d.get("rho_v2_in"),
                dp_overall=d.get("dp_overall"),
                dp_friction=d.get("dp_friction"),
                dp_elevation=d.get("dp_elevation"),
                pressure_in=d.get("pressure_in"),
                pressure_out=d.get("pressure_out"),
                temperature_in=d.get("temperature_in"),
                temperature_out=d.get("temperature_out"),
                density_in=d.get("density_in"),
                density_out=d.get("density_out"),
                viscosity=d.get("viscosity"),
                mass_flow=d.get("mass_flow"),
                vol_flow=d.get("vol_flow"),
                dp_length_criteria_max=d.get("dp_length_criteria_max"),
                velocity_criteria_max=d.get("velocity_criteria_max"),
                velocity_criteria_min=d.get("velocity_criteria_min"),
                units=u,
            )
        )
    return results


# ── MAIN PARSER ───────────────────────────────────────────────────────


def parse_korf_excel(excel_path: str | Path) -> dict[CaseInfo, KorfCaseData]:
    """Parse a KORF Excel report file and extract data for all cases.

    Args:
        excel_path: Path to the KORF Excel report file (.xlsx).

    Returns:
        Dict mapping CaseInfo to KorfCaseData for each case found.
    """
    excel_path = Path(excel_path)
    wb = load_workbook(str(excel_path), data_only=True)

    cases = _parse_cases(wb.sheetnames)
    if not cases:
        _logger.warning("No case sheets found in %s", excel_path)
        return {}

    result: dict[CaseInfo, KorfCaseData] = {}

    for (case_num, case_name), sheets in cases.items():
        case_info = CaseInfo(number=case_num, name=case_name)
        case_data = KorfCaseData(case=case_info)

        # Parse Title sheet
        title_sheet_name = sheets.get("Title")
        if title_sheet_name and title_sheet_name in wb.sheetnames:
            ws = wb[title_sheet_name]
            validations, run_message = _parse_title_sheet(ws)
            case_data.validations = validations
            case_data.run_message = run_message

        # Parse Piping sheet
        piping_sheet_name = sheets.get("Piping")
        if piping_sheet_name and piping_sheet_name in wb.sheetnames:
            ws = wb[piping_sheet_name]
            case_data.pipes = _parse_piping_sheet(ws)

        # Parse Equipment sheet
        equip_sheet_name = sheets.get("Equipment")
        if equip_sheet_name and equip_sheet_name in wb.sheetnames:
            ws = wb[equip_sheet_name]
            equip_data = _parse_equipment_sheet(ws)
            case_data.feeds = equip_data.get("feeds", [])
            case_data.products = equip_data.get("products", [])
            case_data.orifices = equip_data.get("orifices", [])
            case_data.valves = equip_data.get("valves", [])
            case_data.pumps = equip_data.get("pumps", [])
            case_data.compressors = equip_data.get("compressors", [])
            case_data.exchangers = equip_data.get("exchangers", [])
            case_data.misc_equipment = equip_data.get("misc_equipment", [])

        result[case_info] = case_data

    wb.close()
    return result
