from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class CaseInfo:
    """Identifies a single case in the KORF Excel report."""

    number: str
    name: str


@dataclass
class PipeData:
    """Result data for a single pipe, parsed from a Piping sheet."""

    name: str
    length: float | None = None
    size: str | None = None
    schedule: str | None = None
    dp_length: float | None = None
    velocity_in: float | None = None
    rho_v2_in: float | None = None
    dp_overall: float | None = None
    dp_friction: float | None = None
    dp_elevation: float | None = None
    pressure_in: float | None = None
    pressure_out: float | None = None
    temperature_in: float | None = None
    temperature_out: float | None = None
    density_in: float | None = None
    density_out: float | None = None
    viscosity: float | None = None
    mass_flow: float | None = None
    vol_flow: float | None = None
    dp_length_criteria_max: float | None = None
    velocity_criteria_max: float | None = None
    velocity_criteria_min: float | None = None


@dataclass
class FeedData:
    """Result data for a single feed element from Equipment sheet."""

    name: str
    description: str = ""
    pressure: float | None = None


@dataclass
class ProductData:
    """Result data for a single product element from Equipment sheet."""

    name: str
    description: str = ""
    pressure: float | None = None


@dataclass
class OrificeData:
    """Result data for a single orifice from Equipment sheet."""

    name: str
    description: str = ""
    type: str = ""
    bore: float | None = None
    beta: float | None = None
    dp_flange_tap: float | None = None
    dp_pipe_tap: float | None = None
    pressure_in: float | None = None
    pressure_out: float | None = None
    pipe_inlet: str = ""
    pipe_outlet: str = ""


@dataclass
class ValveData:
    """Result data for a single valve from Equipment sheet."""

    name: str
    description: str = ""
    type: str = ""
    cv: float | None = None
    lift_pct: float | None = None
    dp: float | None = None
    pressure_in: float | None = None
    pressure_out: float | None = None
    pipe_inlet: str = ""
    pipe_outlet: str = ""


@dataclass
class PumpData:
    """Result data for a single pump from Equipment sheet."""

    name: str
    description: str = ""
    elevation: float | None = None
    efficiency: float | None = None
    power: float | None = None
    flow: float | None = None
    density: float | None = None
    vol_flow: float | None = None
    head: float | None = None
    dp: float | None = None
    pressure_in: float | None = None
    pressure_out: float | None = None
    pipe_inlet: str = ""
    pipe_outlet: str = ""
    npsha: float | None = None
    npshr: float | None = None
    shutoff_dp: float | None = None
    shutoff_pressure: float | None = None
    suction_max_pressure: float | None = None
    vessel_pressure: float | None = None
    vessel_max_level: float | None = None


@dataclass
class ValidationEntry:
    """A single validation issue from the Title sheet."""

    message: str


@dataclass
class KorfCaseData:
    """All parsed data for a single case from KORF Excel report."""

    case: CaseInfo
    pipes: list[PipeData] = field(default_factory=list)
    feeds: list[FeedData] = field(default_factory=list)
    products: list[ProductData] = field(default_factory=list)
    orifices: list[OrificeData] = field(default_factory=list)
    valves: list[ValveData] = field(default_factory=list)
    pumps: list[PumpData] = field(default_factory=list)
    validations: list[ValidationEntry] = field(default_factory=list)
    run_message: str = ""


_SHEET_PATTERN = re.compile(r"^(Title|Profile|Piping|Equipment)-(\d+)\s+(.+)$")

_SECTION_MARKERS_EQUIPMENT = {
    "FEEDS",
    "PRODUCTS",
    "ORIFICES",
    "VALVES",
    "PUMPS",
    "NPSH",
    "SHUT OFF PRESSURE",
    "CURVES",
    "COMPRESSORS",
    "HEAT EXCHANGERS",
    "MISC EQUIPMENT",
    "JUNCTIONS",
}


def _parse_cases(ws_names: list[str]) -> dict[tuple[str, str], dict[str, str]]:
    """Parse sheet names to discover cases.

    Returns:
        Dict mapping (case_number, case_name) to dict of sheet_type -> sheet_name.
    """
    cases: dict[tuple[str, str], dict[str, str]] = {}
    for name in ws_names:
        m = _SHEET_PATTERN.match(name)
        if m:
            sheet_type, case_num, case_name = m.groups()
            key = (case_num, case_name)
            cases.setdefault(key, {})[sheet_type] = name
    return cases


def _safe_float(val: Any) -> float | None:
    """Convert a cell value to float, returning None if not numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_str(val: Any) -> str:
    """Convert a cell value to string, empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


# ── TITLE SHEET PARSER ──────────────────────────────────────────────


def _parse_title_sheet(ws: Worksheet) -> tuple[list[ValidationEntry], str]:
    """Parse Title sheet to extract validation issues and run message.

    Returns:
        Tuple of (list of ValidationEntry, run_message).
    """
    validations: list[ValidationEntry] = []
    run_message = ""
    in_warnings = False

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value

        if cell_a and "WARNINGS AND ERRORS" in str(cell_a).upper():
            in_warnings = True
            continue

        if in_warnings and cell_a:
            msg = str(cell_a).strip()
            if msg:
                validations.append(ValidationEntry(message=msg))
            continue

        if cell_a and str(cell_a).strip() == "Run Message":
            run_message = str(cell_b).strip() if cell_b else ""

    return validations, run_message


# ── PIPING SHEET PARSER ──────────────────────────────────────────────


def _parse_piping_sheet(ws: Worksheet) -> list[PipeData]:
    """Parse a Piping sheet to extract pipe result data.

    The Piping sheet has a predictable structure:
    - Row 1: 'PIPES' section marker
    - Row 2: 'Number' header with pipe names in columns E onwards
    - Rows 3+: data sections identified by column A labels
    """
    pipe_names: list[str] = []
    pipe_data: dict[str, dict[str, float | None | str]] = {}

    for col in range(5, ws.max_column + 1):
        name = _safe_str(ws.cell(row=2, column=col).value)
        if name and name != "Number":
            pipe_names.append(name)
            pipe_data[name] = {}

    if not pipe_names:
        return []

    def _set_all_pipes(label: str, row_idx: int, sub_key: str | None = None):
        for i, pname in enumerate(pipe_names):
            col = 5 + i
            val = _safe_float(ws.cell(row=row_idx, column=col).value)
            key = f"{label}" if not sub_key else f"{label}_{sub_key}"
            pipe_data[pname][key] = val

    for row_idx in range(1, ws.max_row + 1):
        col_a = _safe_str(ws.cell(row=row_idx, column=1).value)
        col_b = _safe_str(ws.cell(row=row_idx, column=2).value)
        col_c = _safe_str(ws.cell(row=row_idx, column=3).value)
        col_d = _safe_str(ws.cell(row=row_idx, column=4).value)

        if col_a == "Total" and col_b == "Flow":
            _set_all_pipes("mass_flow", row_idx)
        elif col_b == "Pressure" and col_c == "In":
            _set_all_pipes("pressure_in", row_idx)
        elif col_b == "Pressure" and col_c == "Out":
            _set_all_pipes("pressure_out", row_idx)
        elif col_b == "Temperature" and col_c == "In":
            _set_all_pipes("temperature_in", row_idx)
        elif col_b == "Temperature" and col_c == "Out":
            _set_all_pipes("temperature_out", row_idx)
        elif col_b == "Density (No slip)" and col_c == "In":
            _set_all_pipes("density_in", row_idx)
        elif col_b == "Density (No slip)" and col_c == "Out":
            _set_all_pipes("density_out", row_idx)
        elif col_b == "Viscocity (No slip)" or col_b == "Viscosity (No slip)":
            _set_all_pipes("viscosity", row_idx)
        elif col_b == "Volumetric Flow" and col_c == "Avg":
            _set_all_pipes("vol_flow", row_idx)
        elif col_a == "Size" and col_d == "in":
            for i, pname in enumerate(pipe_names):
                pipe_data[pname]["size"] = _safe_str(ws.cell(row=row_idx, column=5 + i).value)
        elif col_a == "Length" and col_d == "m":
            _set_all_pipes("length", row_idx)
        elif col_a == "Schedule":
            for i, pname in enumerate(pipe_names):
                pipe_data[pname]["schedule"] = _safe_str(ws.cell(row=row_idx, column=5 + i).value)
        elif col_a == "dP/Length" and col_b == "" and col_d == "bar/100m":
            _set_all_pipes("dp_length", row_idx)
        elif col_a == "dP/Length" and col_b == "Target" and col_c == "Max":
            _set_all_pipes("dp_length_criteria_max", row_idx)
        elif col_a == "Velocity" and col_b == "" and col_c == "In":
            _set_all_pipes("velocity_in", row_idx)
        elif col_a == "Rho.V^2" or col_a == "Rho.V²" or col_a == "ρV²":  # noqa: RUF001
            if col_c == "In":
                _set_all_pipes("rho_v2_in", row_idx)
        elif col_a == "Overall" and col_d == "bar":
            _set_all_pipes("dp_overall", row_idx)
        elif col_a == "Friction" and col_d == "bar":
            _set_all_pipes("dp_friction", row_idx)
        elif col_a == "Elevation" and col_d == "bar":
            _set_all_pipes("dp_elevation", row_idx)
        elif col_a == "dP/Length" and col_b == "Target" and col_c == "Max":
            _set_all_pipes("dp_length_criteria_max", row_idx)
        elif col_a == "Velocity" and col_b == "Target" and col_c == "Max":
            _set_all_pipes("velocity_criteria_max", row_idx)
        elif col_a == "" and col_b == "" and col_c == "Min" and col_d == "m/s":
            if row_idx > 1:
                row_above_a = _safe_str(ws.cell(row=row_idx - 1, column=1).value)
                row_above_b = _safe_str(ws.cell(row=row_idx - 1, column=2).value)
                if row_above_a == "Velocity" and row_above_b == "Target":
                    _set_all_pipes("velocity_criteria_min", row_idx)

    results: list[PipeData] = []
    for pname in pipe_names:
        d = pipe_data.get(pname, {})
        results.append(
            PipeData(
                name=pname,
                length=d.get("length"),
                size=d.get("size"),
                schedule=d.get("schedule"),
                dp_length=d.get("dp_length"),
                velocity_in=d.get("velocity_in"),
                rho_v2_in=d.get("rho_v2_in"),
                dp_overall=d.get("dp_overall"),
                dp_friction=d.get("dp_friction"),
                dp_elevation=d.get("dp_elevation"),
                pressure_in=d.get("pressure_in"),
                pressure_out=d.get("pressure_out"),
                temperature_in=d.get("temperature_in"),
                temperature_out=d.get("temperature_out"),
                density_in=d.get("density_in"),
                density_out=d.get("density_out"),
                viscosity=d.get("viscosity"),
                mass_flow=d.get("mass_flow"),
                vol_flow=d.get("vol_flow"),
                dp_length_criteria_max=d.get("dp_length_criteria_max"),
                velocity_criteria_max=d.get("velocity_criteria_max"),
                velocity_criteria_min=d.get("velocity_criteria_min"),
            )
        )
    return results


# ── EQUIPMENT SHEET PARSER ───────────────────────────────────────────


def _parse_equipment_sheet(ws: Worksheet) -> dict[str, list]:
    """Parse an Equipment sheet to extract all element data.

    Returns:
        Dict with keys 'feeds', 'products', 'orifices', 'valves', 'pumps',
        each containing a list of parsed dataclass instances.
    """
    result: dict[str, list] = {
        "feeds": [],
        "products": [],
        "orifices": [],
        "valves": [],
        "pumps": [],
    }

    row = 1
    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)

        if cell_a == "FEEDS":
            feeds, row = _parse_feeds_section(ws, row)
            result["feeds"] = feeds
            continue
        elif cell_a == "PRODUCTS":
            products, row = _parse_products_section(ws, row)
            result["products"] = products
            continue
        elif cell_a == "ORIFICES":
            orifices, row = _parse_orifices_section(ws, row)
            result["orifices"] = orifices
            continue
        elif cell_a == "VALVES":
            valves, row = _parse_valves_section(ws, row)
            result["valves"] = valves
            continue
        elif cell_a == "PUMPS":
            pumps, row = _parse_pumps_section(ws, row)
            result["pumps"] = pumps
            continue

        row += 1

    return result


def _read_row(ws: Worksheet, row: int, max_col: int = 25) -> list[Any]:
    """Read all cell values from a row up to max_col."""
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def _find_data_start(ws: Worksheet, start_row: int, marker_col: int = 1) -> int:
    """Find the first data row after headers (headers end when unit row has values)."""
    row = start_row
    for _ in range(10):
        row += 1
        if row > ws.max_row:
            break
        first_val = ws.cell(row=row, column=1).value
        if first_val is not None and str(first_val).strip():
            # Check if this looks like data (has a name/number in col A)
            # and previous row was a unit row
            return row
    return start_row + 4


def _parse_feeds_section(ws: Worksheet, start_row: int) -> tuple[list[FeedData], int]:
    """Parse FEEDS section from Equipment sheet.

    Structure (3 header rows + data rows):
    Row 1: FEEDS marker
    Row 2: Number, Description, Vessel, (empty), Fluid, (empty), Nozzle, ...
    Row 3: (empty), (empty), Pressure, Elevation..., ...
    Row 4: (empty), (empty), barg, m, m, kg/m3, m, bar, barg, bar, bar, barg
    Row 5+: data rows
    End: next section marker in col A
    """
    header_row2 = _read_row(ws, start_row + 1)
    header_row3 = _read_row(ws, start_row + 2)
    header_row4 = _read_row(ws, start_row + 3)

    feeds: list[FeedData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a:
            row += 1
            continue

        name = cell_a
        description = _safe_str(ws.cell(row=row, column=2).value)
        pressure = _safe_float(ws.cell(row=row, column=12).value)

        feeds.append(FeedData(name=name, description=description, pressure=pressure))
        row += 1

    return feeds, row


def _parse_products_section(ws: Worksheet, start_row: int) -> tuple[list[ProductData], int]:
    """Parse PRODUCTS section from Equipment sheet."""
    products: list[ProductData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a:
            row += 1
            continue

        name = cell_a
        description = _safe_str(ws.cell(row=row, column=2).value)
        pressure = _safe_float(ws.cell(row=row, column=12).value)

        products.append(ProductData(name=name, description=description, pressure=pressure))
        row += 1

    return products, row


def _parse_orifices_section(ws: Worksheet, start_row: int) -> tuple[list[OrificeData], int]:
    """Parse ORIFICES section from Equipment sheet.

    Structure (3 header rows + data rows):
    Row 1: ORIFICES marker
    Row 2: Number, Description, Type, Elevation, No Holes, Bore, Beta, Y, C, Pressures, ...
    Row 3: (sub-headers)
    Row 4: (unit row)
    Row 5+: data rows
    """
    orifices: list[OrificeData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a:
            row += 1
            continue

        name = cell_a
        description = _safe_str(ws.cell(row=row, column=2).value)
        orifice_type = _safe_str(ws.cell(row=row, column=3).value)
        bore = _safe_float(ws.cell(row=row, column=6).value)
        beta = _safe_float(ws.cell(row=row, column=7).value)
        dp_flange = _safe_float(ws.cell(row=row, column=10).value)
        dp_pipe = _safe_float(ws.cell(row=row, column=11).value)
        pressure_in = _safe_float(ws.cell(row=row, column=12).value)
        pressure_out = _safe_float(ws.cell(row=row, column=13).value)
        pipe_inlet = _safe_str(ws.cell(row=row, column=14).value)
        pipe_outlet = _safe_str(ws.cell(row=row, column=15).value)

        orifices.append(
            OrificeData(
                name=name,
                description=description,
                type=orifice_type,
                bore=bore,
                beta=beta,
                dp_flange_tap=dp_flange,
                dp_pipe_tap=dp_pipe,
                pressure_in=pressure_in,
                pressure_out=pressure_out,
                pipe_inlet=pipe_inlet,
                pipe_outlet=pipe_outlet,
            )
        )
        row += 1

    return orifices, row


def _parse_valves_section(ws: Worksheet, start_row: int) -> tuple[list[ValveData], int]:
    """Parse VALVES section from Equipment sheet.

    Structure (3 header rows + data rows).
    Key fields: Name col1, Description col2, Type col3, Cv col6,
    DP(Pin-Pout) col13, Inlet Pressure col14, Outlet Pressure col15,
    Pipe Inlet col18, Pipe Outlet col19.
    """
    valves: list[ValveData] = []
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        if not cell_a:
            row += 1
            continue

        name = cell_a
        description = _safe_str(ws.cell(row=row, column=2).value)
        valve_type = _safe_str(ws.cell(row=row, column=3).value)
        cv = _safe_float(ws.cell(row=row, column=6).value)
        dp = _safe_float(ws.cell(row=row, column=13).value)
        pressure_in = _safe_float(ws.cell(row=row, column=14).value)
        pressure_out = _safe_float(ws.cell(row=row, column=15).value)
        pipe_inlet = _safe_str(ws.cell(row=row, column=18).value)
        pipe_outlet = _safe_str(ws.cell(row=row, column=19).value)

        valves.append(
            ValveData(
                name=name,
                description=description,
                type=valve_type,
                cv=cv,
                dp=dp,
                pressure_in=pressure_in,
                pressure_out=pressure_out,
                pipe_inlet=pipe_inlet,
                pipe_outlet=pipe_outlet,
            )
        )
        row += 1

    return valves, row


def _parse_pumps_section(ws: Worksheet, start_row: int) -> tuple[list[PumpData], int]:
    """Parse PUMPS section from Equipment sheet, including NPSH and SHUT OFF PRESSURE.

    Pumps section has PUMPS header, then data, then NPSH section, then
    SHUT OFF PRESSURE section, then optionally CURVES section.
    """
    pumps: list[PumpData] = []

    # Parse PUMPS data (3 header rows + data rows)
    row = start_row + 4

    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if (
            cell_a in ("NPSH", "SHUT OFF PRESSURE", "CURVES")
            or cell_a in _SECTION_MARKERS_EQUIPMENT
        ):
            break
        if not cell_a:
            row += 1
            continue

        name = cell_a
        description = _safe_str(ws.cell(row=row, column=2).value)
        elevation = _safe_float(ws.cell(row=row, column=3).value)
        efficiency = _safe_float(ws.cell(row=row, column=4).value)
        power = _safe_float(ws.cell(row=row, column=5).value)
        flow = _safe_float(ws.cell(row=row, column=6).value)
        density = _safe_float(ws.cell(row=row, column=7).value)
        vol_flow = _safe_float(ws.cell(row=row, column=8).value)
        head = _safe_float(ws.cell(row=row, column=9).value)
        dp = _safe_float(ws.cell(row=row, column=10).value)
        pressure_in = _safe_float(ws.cell(row=row, column=11).value)
        pressure_out = _safe_float(ws.cell(row=row, column=12).value)
        pipe_inlet = _safe_str(ws.cell(row=row, column=13).value)
        pipe_outlet = _safe_str(ws.cell(row=row, column=14).value)

        pumps.append(
            PumpData(
                name=name,
                description=description,
                elevation=elevation,
                efficiency=efficiency,
                power=power,
                flow=flow,
                density=density,
                vol_flow=vol_flow,
                head=head,
                dp=dp,
                pressure_in=pressure_in,
                pressure_out=pressure_out,
                pipe_inlet=pipe_inlet,
                pipe_outlet=pipe_outlet,
            )
        )
        row += 1

    # Parse NPSH section
    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if cell_a == "NPSH":
            # Skip NPSH header rows (3 rows: marker, sub-header, units)
            npsh_data_row = row + 4
            if npsh_data_row <= ws.max_row:
                for pump in pumps:
                    pump_name_in_npsh = _safe_str(ws.cell(row=npsh_data_row, column=1).value)
                    if pump_name_in_npsh == pump.name:
                        pump.npsha = _safe_float(ws.cell(row=npsh_data_row, column=15).value)
                        pump.npshr = _safe_float(ws.cell(row=npsh_data_row, column=18).value)
            row += 1
            continue
        if cell_a in ("SHUT OFF PRESSURE", "CURVES") or cell_a in _SECTION_MARKERS_EQUIPMENT:
            break
        row += 1

    # Parse SHUT OFF PRESSURE section
    while row <= ws.max_row:
        cell_a = _safe_str(ws.cell(row=row, column=1).value)
        if cell_a == "SHUT OFF PRESSURE":
            # Skip header rows
            shutoff_data_row = row + 4
            if shutoff_data_row <= ws.max_row:
                for pump in pumps:
                    pump_name_in_shutoff = _safe_str(ws.cell(row=shutoff_data_row, column=1).value)
                    if pump_name_in_shutoff == pump.name:
                        pump.shutoff_dp = _safe_float(
                            ws.cell(row=shutoff_data_row, column=13).value
                        )
                        pump.shutoff_pressure = _safe_float(
                            ws.cell(row=shutoff_data_row, column=14).value
                        )
                        pump.suction_max_pressure = _safe_float(
                            ws.cell(row=shutoff_data_row, column=9).value
                        )
                        pump.vessel_pressure = _safe_float(
                            ws.cell(row=shutoff_data_row, column=3).value
                        )
                        pump.vessel_max_level = _safe_float(
                            ws.cell(row=shutoff_data_row, column=6).value
                        )
            row += 1
            continue
        if cell_a in ("CURVES",) or (
            cell_a in _SECTION_MARKERS_EQUIPMENT and cell_a != "SHUT OFF PRESSURE"
        ):
            break
        row += 1

    return pumps, row


# ── MAIN PARSER ───────────────────────────────────────────────────────


def parse_korf_excel(excel_path: str | Path) -> dict[CaseInfo, KorfCaseData]:
    """Parse a KORF Excel report file and extract data for all cases.

    Args:
        excel_path: Path to the KORF Excel report file (.xlsx).

    Returns:
        Dict mapping CaseInfo to KorfCaseData for each case found.
    """
    excel_path = Path(excel_path)
    wb = load_workbook(str(excel_path), data_only=True)

    cases = _parse_cases(wb.sheetnames)
    if not cases:
        _logger.warning("No case sheets found in %s", excel_path)
        return {}

    result: dict[CaseInfo, KorfCaseData] = {}

    for (case_num, case_name), sheets in cases.items():
        case_info = CaseInfo(number=case_num, name=case_name)
        case_data = KorfCaseData(case=case_info)

        # Parse Title sheet
        title_sheet_name = sheets.get("Title")
        if title_sheet_name and title_sheet_name in wb.sheetnames:
            ws = wb[title_sheet_name]
            validations, run_message = _parse_title_sheet(ws)
            case_data.validations = validations
            case_data.run_message = run_message

        # Parse Piping sheet
        piping_sheet_name = sheets.get("Piping")
        if piping_sheet_name and piping_sheet_name in wb.sheetnames:
            ws = wb[piping_sheet_name]
            case_data.pipes = _parse_piping_sheet(ws)

        # Parse Equipment sheet
        equip_sheet_name = sheets.get("Equipment")
        if equip_sheet_name and equip_sheet_name in wb.sheetnames:
            ws = wb[equip_sheet_name]
            equip_data = _parse_equipment_sheet(ws)
            case_data.feeds = equip_data.get("feeds", [])
            case_data.products = equip_data.get("products", [])
            case_data.orifices = equip_data.get("orifices", [])
            case_data.valves = equip_data.get("valves", [])
            case_data.pumps = equip_data.get("pumps", [])

        result[case_info] = case_data

    wb.close()
    return result
