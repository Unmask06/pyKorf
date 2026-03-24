"""Batch report generation for multiple KDF files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from pykorf import Model
from pykorf.log import get_logger
from pykorf.reports.exporter import ResultExporter

logger = get_logger()


class BatchReportGenerator:
    """Generate comprehensive Excel reports from multiple KDF files.

    Processes all .kdf files in a folder and creates a single Excel workbook
    with multiple sheets (one per element type), each containing elements
    from all files with source_file metadata.

    Example:
        >>> generator = BatchReportGenerator("/path/to/kdf/folder")
        >>> generator.generate_report("output.xlsx")
        '/path/to/kdf/folder/output.xlsx'
    """

    ELEMENT_TYPES = ["Pipes", "Products", "Feeds", "Pumps", "Compressors", "Valves"]

    def __init__(self, folder_path: str | Path):
        """Initialize batch report generator.

        Args:
            folder_path: Path to folder containing KDF files.
        """
        self.folder = Path(folder_path)
        self.kdf_files: list[Path] = []
        self._errors: list[str] = []

    @property
    def file_count(self) -> int:
        """Return number of KDF files discovered."""
        return len(self.kdf_files)

    @property
    def errors(self) -> list[str]:
        """Return list of errors encountered during processing."""
        return self._errors.copy()

    def discover_files(self) -> int:
        """Find all .kdf files in folder.

        Returns:
            Number of KDF files found.

        Raises:
            ValueError: If folder doesn't exist or isn't a directory.
        """
        if not self.folder.exists():
            raise ValueError(f"Folder not found: {self.folder}")
        if not self.folder.is_dir():
            raise ValueError(f"Not a folder: {self.folder}")

        self.kdf_files = sorted(self.folder.rglob("*.kdf"))
        logger.info("batch_discover", folder=str(self.folder), count=len(self.kdf_files))
        return len(self.kdf_files)

    def generate_report(
        self,
        output_path: str | Path | None = None,
        element_types: list[str] | None = None,
        include_line_numbers: bool = True,
    ) -> str:
        """Generate multi-sheet Excel report.

        Args:
            output_path: Destination .xlsx file. If None, creates in source folder.
            element_types: Filter by types ['Pipes', 'Pumps', ...]. None = all.
            include_line_numbers: Parse and add line_number column for pipes.

        Returns:
            Path to generated Excel file.

        Raises:
            ValueError: If no KDF files found or folder invalid.
        """
        if not self.kdf_files:
            self.discover_files()

        if not self.kdf_files:
            raise ValueError(f"No KDF files found in: {self.folder}")

        if output_path is None:
            output_path = self.folder / f"{self.folder.name}_batch_report.xlsx"
        else:
            output_path = Path(output_path)

        types_to_process = element_types or self.ELEMENT_TYPES

        elements_by_type: dict[str, list[dict[str, Any]]] = {
            etype: [] for etype in self.ELEMENT_TYPES
        }

        logger.info("batch_report_start", folder=str(self.folder), files=len(self.kdf_files))
        for i, kdf_file in enumerate(self.kdf_files, 1):
            logger.info(
                "batch_report_processing", index=i, total=len(self.kdf_files), file=kdf_file.name
            )
            try:
                model = Model.load(str(kdf_file))
                exporter = ResultExporter(model)

                for sheet_name in types_to_process:
                    if sheet_name not in exporter._extractors:
                        continue

                    extractor = exporter._extractors[sheet_name]
                    elements = extractor()

                    for elem in elements:
                        elem["source_file"] = kdf_file.name
                        elem["source_path"] = str(kdf_file)

                    elements_by_type[sheet_name].extend(elements)

                logger.info(
                    "batch_report_done", index=i, total=len(self.kdf_files), file=kdf_file.name
                )
            except Exception as e:
                error_msg = f"Error processing {kdf_file.name}: {e}"
                self._errors.append(error_msg)
                logger.error("batch_process_error", file=str(kdf_file), error=str(e))

        self._write_excel(output_path, elements_by_type)

        logger.info(
            "batch_report_complete",
            output=str(output_path),
            files_processed=len(self.kdf_files),
            errors=len(self._errors),
        )

        return str(output_path)

    def _write_excel(self, output_path: Path, elements_by_type: dict[str, list[dict]]) -> None:
        """Write elements to Excel with multiple sheets, preserving existing custom sheets.

        If the Excel file exists, non-element sheets (e.g., Dashboard, Summary) are preserved.
        Element sheets are always recreated with fresh data. If file is locked or corrupted,
        creates a new file with _v2, _v3 suffix.

        Args:
            output_path: Destination Excel file path.
            elements_by_type: Dict mapping sheet names to element lists.
        """
        workbook: openpyxl.Workbook | None = None
        output_path_actual = output_path

        if output_path.exists():
            try:
                workbook = openpyxl.load_workbook(output_path)

                for sheet_name in self.ELEMENT_TYPES:
                    if sheet_name in workbook.sheetnames:
                        del workbook[sheet_name]

                logger.info(
                    "batch_report_preserve",
                    path=str(output_path),
                    preserved_sheets=[
                        s for s in workbook.sheetnames if s not in self.ELEMENT_TYPES
                    ],
                )

            except (PermissionError, OSError) as e:
                logger.warning(
                    "batch_report_file_locked",
                    path=str(output_path),
                    error=str(e),
                )
                output_path_actual = self._create_unique_path(output_path)
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)

            except Exception as e:
                logger.warning(
                    "batch_report_corrupted",
                    path=str(output_path),
                    error=str(e),
                )
                output_path_actual = self._create_unique_path(output_path)
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

        for sheet_name, elements in elements_by_type.items():
            if not elements:
                continue

            df = pd.DataFrame(elements)

            cols = list(df.columns)
            for col in ["source_path", "source_file"]:
                if col in cols:
                    cols.remove(col)
            cols = ["source_file", *cols]

            if sheet_name == "Pipes" and "Line Number" in df.columns:
                cols.remove("Line Number")
                cols.insert(1, "Line Number")

            if "source_path" in df.columns:
                cols.append("source_path")

            df = df[[c for c in cols if c in df.columns]]

            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                ws.delete_rows(1, ws.max_row)
            else:
                ws = workbook.create_sheet(sheet_name)

            for c_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=c_idx, value=col_name).font = openpyxl.styles.Font(bold=True)

            for r_idx, row in enumerate(df.values, start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column].width = adjusted_width

        workbook.save(output_path_actual)

        logger.info(
            "batch_report_saved",
            path=str(output_path_actual),
            sheets_written=len(elements_by_type),
        )

    def _create_unique_path(self, base_path: Path) -> Path:
        """Create a unique file path by appending _v2, _v3, etc.

        Args:
            base_path: Original desired path.

        Returns:
            Unique path that doesn't exist or is not locked.
        """
        counter = 2
        while True:
            new_path = base_path.with_stem(f"{base_path.stem}_v{counter}")
            if not new_path.exists():
                try:
                    test_wb = openpyxl.Workbook()
                    test_wb.save(new_path)
                    test_wb.close()
                    new_path.unlink(missing_ok=True)
                    return new_path
                except (PermissionError, OSError):
                    counter += 1
            else:
                counter += 1

    def get_preserved_sheets(self, output_path: Path) -> list[str]:
        """Return list of non-element sheets that will be preserved.

        Args:
            output_path: Path to existing Excel file.

        Returns:
            List of sheet names that will be kept (excludes element sheets).
        """
        if not output_path.exists():
            return []

        try:
            wb = openpyxl.load_workbook(output_path, read_only=True)
            sheets = set(wb.sheetnames)
            wb.close()

            return list(sheets - set(self.ELEMENT_TYPES))
        except Exception:
            return []
