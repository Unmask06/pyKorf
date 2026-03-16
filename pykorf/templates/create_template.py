"""Create default Excel template for pyKorf reports."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def create_template():
    """Create A3 landscape Excel template with logo placeholders."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Set A3 landscape page setup
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Set up logo placeholders (rows 1-5)
    ws.cell(row=1, column=1, value="[Company Logo]")
    ws.cell(row=1, column=1).font = Font(italic=True, color="808080")

    ws.cell(row=1, column=10, value="[Contractor Logo]")
    ws.cell(row=1, column=10).font = Font(italic=True, color="808080")

    # Set column widths
    column_widths = {
        "A": 5,
        "B": 15,
        "C": 15,
        "D": 20,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 15,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save template
    template_path = Path(__file__).parent / "report_template.xlsx"
    wb.save(template_path)
    print(f"Template created: {template_path}")


if __name__ == "__main__":
    create_template()
