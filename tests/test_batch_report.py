"""Tests for batch report sheet preservation functionality."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from pykorf.use_case.batch_report import BatchReportGenerator


class TestBatchReportSheetPreservation:
    """Test that custom sheets are preserved when regenerating batch reports."""

    def test_get_preserved_sheets_empty_path(self):
        """Return empty list when file doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = BatchReportGenerator(tmpdir)
            result = generator.get_preserved_sheets(Path(tmpdir) / "nonexistent.xlsx")
            assert result == []

    def test_get_preserved_sheets_all_custom(self):
        """All non-element sheets should be preserved."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            # Create Excel with only custom sheets (remove default sheet)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default 'Sheet'
            wb.create_sheet("Dashboard")
            wb.create_sheet("Summary")
            wb.create_sheet("Notes")
            test_excel = tmpdir / "test.xlsx"
            wb.save(test_excel)
            wb.close()

            generator = BatchReportGenerator(tmpdir)
            preserved = generator.get_preserved_sheets(test_excel)

            assert "Dashboard" in preserved
            assert "Summary" in preserved
            assert "Notes" in preserved
            assert len(preserved) == 3

    def test_get_preserved_sheets_mixed(self):
        """Element sheets should be excluded from preserved list."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default 'Sheet'
            wb.create_sheet("Dashboard")
            wb.create_sheet("Pipes")  # Element sheet
            wb.create_sheet("Pumps")  # Element sheet
            wb.create_sheet("Summary")
            test_excel = tmpdir / "test.xlsx"
            wb.save(test_excel)
            wb.close()

            generator = BatchReportGenerator(tmpdir)
            preserved = generator.get_preserved_sheets(test_excel)

            assert "Dashboard" in preserved
            assert "Summary" in preserved
            assert "Pipes" not in preserved
            assert "Pumps" not in preserved

    def test_create_unique_path(self):
        """Unique path should increment version number."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            base_path = tmpdir / "report.xlsx"
            base_path.touch()  # Create file

            generator = BatchReportGenerator(tmpdir)
            unique_path = generator._create_unique_path(base_path)

            assert unique_path.name == "report_v2.xlsx"
            assert unique_path.exists() is False  # Should not create file, just check availability

    def test_create_unique_path_multiple_versions(self):
        """Should handle multiple existing versions."""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            base_path = tmpdir / "report.xlsx"
            base_path.touch()
            (tmpdir / "report_v2.xlsx").touch()
            (tmpdir / "report_v3.xlsx").touch()

            generator = BatchReportGenerator(tmpdir)
            unique_path = generator._create_unique_path(base_path)

            assert unique_path.name == "report_v4.xlsx"

    def test_write_excel_preserves_custom_sheets(self, tmp_path):
        """Element sheets are replaced, custom sheets preserved."""
        # Create initial Excel with mixed sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default 'Sheet'
        wb.create_sheet("Dashboard")
        wb.create_sheet("Pipes")
        wb.create_sheet("Summary")

        ws_dash = wb["Dashboard"]
        ws_dash["A1"] = "Original Dashboard"

        ws_summary = wb["Summary"]
        ws_summary["A1"] = "Original Summary"

        test_excel = tmp_path / "test_batch_report.xlsx"
        wb.save(test_excel)
        wb.close()

        # Mock element data
        elements_by_type = {
            "Pipes": [{"source_file": "test.kdf", "name": "L1", "diameter": "6"}],
            "Pumps": [],
            "Feeds": [],
            "Products": [],
            "Compressors": [],
            "Valves": [],
        }

        generator = BatchReportGenerator(tmp_path)
        generator._write_excel(test_excel, elements_by_type)

        # Verify
        wb_result = openpyxl.load_workbook(test_excel)

        # Custom sheets preserved
        assert "Dashboard" in wb_result.sheetnames
        assert "Summary" in wb_result.sheetnames
        assert wb_result["Dashboard"]["A1"].value == "Original Dashboard"
        assert wb_result["Summary"]["A1"].value == "Original Summary"

        # Element sheets recreated
        assert "Pipes" in wb_result.sheetnames
        pipes_sheet = wb_result["Pipes"]
        assert pipes_sheet.max_row >= 1  # Has data (at least headers)

    @pytest.mark.skip(reason="File locking behavior is platform-specific; skip on Windows")
    def test_write_excel_locked_file_creates_version(self, tmp_path):
        """Locked file should create versioned copy."""
        test_excel = tmp_path / "test_batch_report.xlsx"

        # Create file with some data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Test Data"
        wb.save(test_excel)
        wb.close()

        # Simulate lock by opening in read-only mode
        wb_locked = openpyxl.load_workbook(test_excel)

        elements_by_type = {
            "Pipes": [{"source_file": "test.kdf", "name": "L1"}],
            "Pumps": [],
            "Feeds": [],
            "Products": [],
            "Compressors": [],
            "Valves": [],
        }

        generator = BatchReportGenerator(tmp_path)

        try:
            generator._write_excel(test_excel, elements_by_type)
        finally:
            wb_locked.close()

        # Should have created versioned file
        versioned = tmp_path / "test_batch_report_v2.xlsx"
        assert versioned.exists(), f"Expected {versioned} to be created"
