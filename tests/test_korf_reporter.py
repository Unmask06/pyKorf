"""Tests for KorfReporter, korf_parser, and multi-case ResultExporter."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from pykorf import Model
from pykorf.core.reports.exporter import ResultExporter
from pykorf.core.reports.korf_parser import (
    CaseInfo,
    OrificeData,
    PipeData,
    PumpData,
    parse_korf_excel,
)
from pykorf.core.reports.korf_reporter import KorfReporter
from pykorf.core.reports.reporter import PykorfReporter, Reporter

SAMPLES_DIR = Path(__file__).parent.parent / "pykorf" / "library"
PUMP_KDF = SAMPLES_DIR / "Pumpcases.kdf"


def _create_korf_excel(path: Path, num_cases: int = 1) -> Path:
    """Create a minimal KORF Excel file for testing.

    Generates a workbook with Title, Piping, and Equipment sheets
    for the specified number of cases.
    """
    wb = openpyxl.Workbook()

    for case_num in range(1, num_cases + 1):
        case_name = "Rated" if case_num == 1 else f"Case{case_num}"
        suffix = f"{case_num} {case_name}"

        # ── Title sheet ─────────────────────────────────────────────
        ws_title = wb.create_sheet(f"Title-{suffix}")
        ws_title.append(["Model Title"])
        ws_title.append(["", ""])
        ws_title.append(["WARNINGS AND ERRORS"])
        ws_title.append(["Pipe 'VCL22-711' exceeds criteria"])

        # ── Piping sheet ─────────────────────────────────────────────
        ws_pipe = wb.create_sheet(f"Piping-{suffix}")
        # Row 1: section marker
        ws_pipe.append(["PIPES"])
        # Row 2: header row with pipe names
        ws_pipe.append(["Number", "", "", "", "VCL22-711", "L1"])
        # Row 3-4: intermediate headers (Name, Fluid per real KORF structure)
        ws_pipe.append(["Name", "", "", "", "Pipe", "Pipe"])
        ws_pipe.append(["Fluid", "", "", "", "", ""])
        # Row 5+: data rows (Col A=field, Col B="", Col C="", Col D=unit per real KORF structure)
        ws_pipe.append(["Size", "", "", "in", "6", "4"])
        ws_pipe.append(["Length", "", "", "m", 75, 50])
        ws_pipe.append(["Schedule", "", "", "", "40", "40"])
        ws_pipe.append(["dP/Length", "", "", "bar/100m", 0.008, 0.01])
        ws_pipe.append(["dP/Length", "Target", "Max", "bar/100m", 0.09, 0.1])
        ws_pipe.append(["Velocity", "", "In", "m/s", 0.614, 0.9])
        ws_pipe.append(["Velocity", "Target", "Max", "m/s", 1.2, 1.5])
        ws_pipe.append(["", "", "Min", "m/s", 0.0, 0.0])
        ws_pipe.append(["", "", "Out", "m/s", 0.6, 0.85])
        ws_pipe.append(["Rho.V^2", "", "In", "Pa", 350, 280])
        ws_pipe.append(["Overall", "", "", "bar", 0.006, 0.005])
        ws_pipe.append(["Friction", "", "", "bar", 0.005, 0.004])
        ws_pipe.append(["Elevation", "", "", "bar", 0.001, 0.001])
        ws_pipe.append(["Total", "Flow", "", "", 234000, 180000])
        ws_pipe.append(["", "Pressure", "In", "barg", 11.36, 10.5])
        ws_pipe.append(["", "Pressure", "Out", "barg", 11.35, 10.49])
        ws_pipe.append(["", "Density (No slip)", "In", "kg/m3", 928, 950])
        ws_pipe.append(["", "Density (No slip)", "Out", "kg/m3", 927, 949])
        ws_pipe.append(["", "Viscosity (No slip)", "", "cP", 0.3, 0.4])
        ws_pipe.append(["", "Volumetric Flow", "Avg", "m3/h", 252, 189])

        # Add extra rows to reach row 20+ (padding)
        for _ in range(5):
            ws_pipe.append([])

        # ── Equipment sheet ──────────────────────────────────────────
        ws_equip = wb.create_sheet(f"Equipment-{suffix}")

        # FEEDS section
        ws_equip.append(["FEEDS"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Vessel",
                "",
                "Fluid",
                "",
                "Nozzle",
                "",
                "",
                "Pipe",
                "",
                "",
                "",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "Pressure",
                "Elevation",
                "Level",
                "Density",
                "Elevation",
                "dP Level",
                "Pressure",
                "dP Velocity",
                "dP Friction",
                "Pressure",
                "Number",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "barg",
                "m",
                "m",
                "kg/m3",
                "m",
                "bar",
                "barg",
                "bar",
                "bar",
                "barg",
                "",
            ]
        )
        ws_equip.append(
            [
                "170-TC-009",
                "VC PRODUCT SPHERE",
                6.27,
                0,
                0,
                928,
                0,
                0,
                6.27,
                0,
                0,
                6.27,
                "EEG-001",
            ]
        )

        # PRODUCTS section
        ws_equip.append(["PRODUCTS"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Vessel",
                "",
                "Fluid",
                "",
                "Nozzle",
                "",
                "",
                "Pipe",
                "",
                "",
                "",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "Pressure",
                "Elevation",
                "Level",
                "Density",
                "Elevation",
                "dP Level",
                "Pressure",
                "dP Velocity",
                "dP Friction",
                "Pressure",
                "Number",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "barg",
                "m",
                "m",
                "kg/m3",
                "m",
                "bar",
                "barg",
                "bar",
                "bar",
                "barg",
                "",
            ]
        )
        ws_equip.append(
            [
                "TLT Export",
                "Product",
                8.5,
                0,
                0,
                928,
                0,
                0,
                8.5,
                0,
                0,
                8.5,
                "RGE02-421",
            ]
        )
        ws_equip.append(
            [
                "PVC Plant",
                "Product",
                10.07,
                0,
                0,
                928,
                0,
                0,
                10.07,
                0,
                0,
                10.07,
                "RGE02-422",
            ]
        )

        # ORIFICES section
        ws_equip.append(["ORIFICES"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Type",
                "Elevation",
                "No Holes",
                "Bore",
                "Beta",
                "",
                "",
                "Pressures",
                "",
                "",
                "",
                "Pipe",
                "",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "dP Flange Tap",
                "dP Pipe Tap",
                "Inlet",
                "Outlet",
                "Inlet",
                "Outlet",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "",
                "m",
                "",
                "mm",
                "",
                "",
                "",
                "bar",
                "bar",
                "barg",
                "barg",
                "",
                "",
            ]
        )
        ws_equip.append(
            [
                "FT-0701",
                "Orifice",
                "Orifice",
                0,
                1,
                140.5,
                0.5,
                "",
                "",
                0.189,
                0.1,
                11.98,
                11.88,
                "VCL17-702",
                "VCL17-702_1",
            ]
        )

        # VALVES section
        ws_equip.append(["VALVES"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Type",
                "Elevation",
                "Characteristics",
                "Cv",
                "",
                "",
                "",
                "",
                "",
                "",
                "Pressures",
                "",
                "",
                "",
                "",
                "Pipe",
                "",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Pin-Pout",
                "Inlet",
                "Outlet",
                "",
                "",
                "Inlet",
                "Outlet",
            ]
        )
        ws_equip.append(
            [
                "",
                "",
                "",
                "m",
                "",
                "Full Open",
                "",
                "",
                "",
                "",
                "",
                "",
                "bar",
                "barg",
                "barg",
                "",
                "",
                "",
                "",
            ]
        )
        ws_equip.append(
            [
                "FV-0714",
                "Control Valve",
                "FCV",
                0,
                "Linear",
                120,
                "",
                "",
                "",
                "",
                "",
                "",
                0.7,
                11.35,
                10.65,
                "",
                "",
                "VCL22-711",
                "VCL22-711_1",
            ]
        )

        # PUMPS section
        ws_equip.append(["PUMPS"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Elevation",
                "Efficiency",
                "Power",
                "Flow",
                "Density",
                "Vol Flow",
                "Head",
                "DP",
                "Pin",
                "Pout",
                "Pipe In",
                "Pipe Out",
            ]
        )
        ws_equip.append(
            ["", "", "m", "%", "kW", "kg/h", "kg/m3", "m3/h", "m", "bar", "barg", "barg", "", ""]
        )
        ws_equip.append(
            ["", "", "m", "%", "kW", "kg/h", "kg/m3", "m3/h", "m", "bar", "barg", "barg", "", ""]
        )
        ws_equip.append(
            [
                "170-PS-009",
                "VC Product Circ Pump",
                0.7,
                75,
                46.5,
                234000,
                928,
                252,
                54,
                4.92,
                7.04,
                11.96,
                "L5",
                "VCL22-711",
            ]
        )

        # NPSH section
        ws_equip.append(["NPSH"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Elevation",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "NPSHa",
                "",
                "",
                "NPSHr",
                "",
            ]
        )
        ws_equip.append(["", "", "m", "", "", "", "", "", "", "", "", "", "", "m", "", "", "m", ""])
        ws_equip.append(["", "", "m", "", "", "", "", "", "", "", "", "", "", "m", "", "", "m", ""])
        npsh_row = [
            "170-PS-009",
            "VC Product Circ Pump",
            0.7,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            9.5,
            "",
            "",
            7.0,
            "",
        ]
        ws_equip.append(npsh_row)

        # SHUT OFF PRESSURE section
        ws_equip.append(["SHUT OFF PRESSURE"])
        ws_equip.append(
            [
                "Number",
                "Description",
                "Vessel P",
                "",
                "",
                "Vessel Max Level",
                "",
                "",
                "",
                "",
                "",
                "",
                "Shut-Off DP",
                "Shut-Off P",
                "",
                "",
                "",
                "",
            ]
        )
        ws_equip.append(
            ["", "", "barg", "", "", "m", "", "", "", "", "", "", "bar", "barg", "", "", "", ""]
        )
        ws_equip.append(
            ["", "", "barg", "", "", "m", "", "", "", "", "", "", "bar", "barg", "", "", "", ""]
        )
        shutoff_row = [
            "170-PS-009",
            "VC Product Circ Pump",
            10.0,
            "",
            "",
            18.9,
            "",
            "",
            "",
            "",
            "",
            "",
            6.15,
            18.65,
            "",
            "",
            "",
            "",
        ]
        ws_equip.append(shutoff_row)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    out_path = path / "test_korf.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path


# ── Parser Tests ──────────────────────────────────────────────────────────


class TestKorfParser:
    """Tests for KORF Excel parser."""

    def test_parse_korf_excel_finds_cases(self, tmp_path):
        """Parser discovers all cases from sheet names."""
        excel_path = _create_korf_excel(tmp_path, num_cases=2)
        result = parse_korf_excel(excel_path)
        assert len(result) == 2

    def test_parse_korf_excel_case_info(self, tmp_path):
        """Parser extracts correct CaseInfo numbers and names."""
        excel_path = _create_korf_excel(tmp_path, num_cases=1)
        result = parse_korf_excel(excel_path)
        cases = list(result.keys())
        assert cases[0].number == "1"
        assert cases[0].name == "Rated"

    def test_parse_korf_excel_pipes(self, tmp_path):
        """Parser extracts pipe data from Piping sheet."""
        excel_path = _create_korf_excel(tmp_path, num_cases=1)
        result = parse_korf_excel(excel_path)
        case_data = list(result.values())[0]
        assert len(case_data.pipes) >= 2

        pipe = next(p for p in case_data.pipes if p.name == "VCL22-711")
        assert pipe.length == 75
        assert pipe.dp_length is not None
        assert pipe.velocity_in is not None

    def test_parse_korf_excel_pumps(self, tmp_path):
        """Parser extracts pump data from Equipment sheet."""
        excel_path = _create_korf_excel(tmp_path, num_cases=1)
        result = parse_korf_excel(excel_path)
        case_data = list(result.values())[0]
        assert len(case_data.pumps) >= 1

        pump = case_data.pumps[0]
        assert pump.name == "170-PS-009"
        assert pump.elevation is not None

    def test_parse_korf_excel_orifices(self, tmp_path):
        """Parser extracts orifice data from Equipment sheet."""
        excel_path = _create_korf_excel(tmp_path, num_cases=1)
        result = parse_korf_excel(excel_path)
        case_data = list(result.values())[0]
        assert len(case_data.orifices) >= 1

        orif = case_data.orifices[0]
        assert orif.name == "FT-0701"
        assert orif.type == "Orifice"
        assert orif.bore is not None

    def test_parse_korf_excel_valves(self, tmp_path):
        """Parser extracts valve data from Equipment sheet."""
        excel_path = _create_korf_excel(tmp_path, num_cases=1)
        result = parse_korf_excel(excel_path)
        case_data = list(result.values())[0]
        assert len(case_data.valves) >= 1

        valve = case_data.valves[0]
        assert valve.name == "FV-0714"
        assert valve.dp is not None

    def test_parse_korf_excel_validations(self, tmp_path):
        """Parser extracts validation issues from Title sheet."""
        excel_path = _create_korf_excel(tmp_path, num_cases=1)
        result = parse_korf_excel(excel_path)
        case_data = list(result.values())[0]
        assert len(case_data.validations) >= 1
        assert "exceeds criteria" in case_data.validations[0].message

    def test_parse_nonexistent_file(self):
        """Parser raises on nonexistent file."""
        with pytest.raises((FileNotFoundError, OSError)):
            parse_korf_excel(Path("/nonexistent/file.xlsx"))


# ── KorfReporter Tests ────────────────────────────────────────────────────


class TestKorfReporter:
    """Tests for KorfReporter data extraction."""

    @pytest.fixture
    def model(self):
        return Model.load(PUMP_KDF)

    @pytest.fixture
    def korf_excel(self, tmp_path):
        return _create_korf_excel(tmp_path, num_cases=2)

    def test_korf_reporter_satisfies_protocol(self, model, korf_excel):
        """KorfReporter implements the Reporter protocol."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        assert isinstance(reporter, Reporter)

    def test_korf_reporter_source_name(self, model, korf_excel):
        """KorfReporter returns the Excel file name."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        assert reporter.get_source_name() == "test_korf.xlsx"

    def test_korf_reporter_case_names(self, model, korf_excel):
        """KorfReporter returns correct case names."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        names = reporter.get_case_names()
        assert len(names) == 2
        assert "1 - Rated" in names

    def test_korf_reporter_generate_dataframes(self, model, korf_excel):
        """KorfReporter generates DataFrames for primary case."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        dfs = reporter.generate_dataframes()
        assert "Pipes" in dfs
        assert "Pumps" in dfs
        assert not dfs["Pipes"].empty

    def test_korf_reporter_all_case_dataframes(self, model, korf_excel):
        """KorfReporter generates per-case DataFrames."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        all_dfs = reporter.generate_all_case_dataframes()
        assert len(all_dfs) == 2
        for case_name, case_dfs in all_dfs.items():
            assert "Pipes" in case_dfs

    def test_korf_reporter_validation_dataframe(self, model, korf_excel):
        """KorfReporter builds validation DataFrame with case prefix."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        val_df = reporter.get_validation_dataframe()
        assert not val_df.empty
        assert "Case" in val_df.columns[0] or "Message" in val_df.columns

    def test_korf_reporter_pipe_criteria_check(self, model, korf_excel):
        """KorfReporter includes Criteria Check column for pipes."""
        reporter = KorfReporter(excel_path=korf_excel, model=model)
        dfs = reporter.generate_dataframes()
        assert "Criteria Check" in dfs["Pipes"].columns

    def test_pykorf_reporter_satisfies_protocol(self, model):
        """PykorfReporter implements the Reporter protocol."""
        reporter = PykorfReporter(model=model)
        assert isinstance(reporter, Reporter)

    def test_pykorf_reporter_all_case_dataframes(self, model):
        """PykorfReporter's generate_all_case_dataframes returns single-case dict."""
        reporter = PykorfReporter(model=model)
        all_dfs = reporter.generate_all_case_dataframes()
        assert len(all_dfs) == 1


# ── ResultExporter Multi-Case Tests ────────────────────────────────────────


class TestResultExporterMultiCase:
    """Tests for multi-case ResultExporter layout."""

    @pytest.fixture
    def model(self):
        return Model.load(PUMP_KDF)

    @pytest.fixture
    def korf_excel(self, tmp_path):
        return _create_korf_excel(tmp_path, num_cases=2)

    def test_multi_case_creates_per_case_sheets(self, model, korf_excel, tmp_path):
        """Multi-case export creates per-case sheets + Summary."""
        from pykorf.core.reports.korf_reporter import KorfReporter

        reporter = KorfReporter(excel_path=korf_excel, model=model)
        exporter = ResultExporter(reporter=reporter)
        out = exporter.export_to_excel(str(tmp_path / "multi_report.xlsx"))

        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "1 - Rated" in wb.sheetnames
        assert "2 - Case2" in wb.sheetnames
        wb.close()

    def test_single_case_creates_model_summary(self, model, tmp_path):
        """Single-case export creates Model Summary sheet."""
        exporter = ResultExporter(model=model)
        out = exporter.export_to_excel(str(tmp_path / "single_report.xlsx"))

        wb = openpyxl.load_workbook(out)
        assert "Model Summary" in wb.sheetnames
        wb.close()

    def test_multi_case_summary_has_pipe_stats(self, model, korf_excel, tmp_path):
        """Summary envelope sheet includes pipe stats (Min-Max, Overall Criteria)."""
        from pykorf.core.reports.korf_reporter import KorfReporter

        reporter = KorfReporter(excel_path=korf_excel, model=model)
        exporter = ResultExporter(reporter=reporter)
        out = exporter.export_to_excel(str(tmp_path / "summary_report.xlsx"))

        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        # Find "Overall Criteria" or "Min - Max" in sheet
        found_stats = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
            for cell in row:
                if cell.value and (
                    "Overall Criteria" in str(cell.value) or "Min - Max" in str(cell.value)
                ):
                    found_stats = True
                    break
        assert found_stats, "Summary sheet should contain pipe stats"
        wb.close()

    def test_korf_excel_auto_detection(self, tmp_path):
        """_find_korf_excel finds .xlsx alongside .kdf when xlsx is newer."""
        import time

        from pykorf.app.api.routers.report import _find_korf_excel

        # Create .kdf first, then .xlsx so xlsx is newer
        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")
        time.sleep(0.05)
        xlsx_path = tmp_path / "Pumpcases.xlsx"
        xlsx_path.write_text("dummy")

        result = _find_korf_excel(kdf_path)
        assert result is not None
        assert result.name == "Pumpcases.xlsx"

    def test_korf_excel_stale_ignored(self, tmp_path):
        """_find_korf_excel returns None when .xlsx is older than .kdf."""
        import time

        from pykorf.app.api.routers.report import _find_korf_excel

        # Create .xlsx first, then .kdf so xlsx is older
        xlsx_path = tmp_path / "Pumpcases.xlsx"
        xlsx_path.write_text("dummy")
        time.sleep(0.05)
        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")

        result = _find_korf_excel(kdf_path)
        assert result is None

    def test_korf_excel_no_detection(self, tmp_path):
        """_find_korf_excel returns None when no .xlsx found."""
        from pykorf.app.api.routers.report import _find_korf_excel

        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")

        result = _find_korf_excel(kdf_path)
        assert result is None

    def test_korf_excel_exact_stem_only(self, tmp_path):
        """_find_korf_excel ignores {stem}_report.xlsx and {stem}_export.xlsx."""
        from pykorf.app.api.routers.report import _find_korf_excel

        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")
        (tmp_path / "Pumpcases_report.xlsx").write_text("dummy")
        (tmp_path / "Pumpcases_export.xlsx").write_text("dummy")

        result = _find_korf_excel(kdf_path)
        assert result is None

    def test_korf_status_fresh(self, tmp_path):
        """_korf_excel_status returns path and is_stale=False when xlsx is newer."""
        import time

        from pykorf.app.api.routers.report import _korf_excel_status

        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")
        time.sleep(0.05)
        xlsx_path = tmp_path / "Pumpcases.xlsx"
        xlsx_path.write_text("dummy")

        result = _korf_excel_status(kdf_path)
        assert result.korf_excel_path is not None
        assert result.is_stale is False

    def test_korf_status_stale(self, tmp_path):
        """_korf_excel_status returns path and is_stale=True when xlsx is older."""
        import time

        from pykorf.app.api.routers.report import _korf_excel_status

        xlsx_path = tmp_path / "Pumpcases.xlsx"
        xlsx_path.write_text("dummy")
        time.sleep(0.05)
        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")

        result = _korf_excel_status(kdf_path)
        assert result.korf_excel_path is not None
        assert result.is_stale is True

    def test_korf_status_no_excel(self, tmp_path):
        """_korf_excel_status returns None path when no xlsx exists."""
        from pykorf.app.api.routers.report import _korf_excel_status

        kdf_path = tmp_path / "Pumpcases.kdf"
        kdf_path.write_text("dummy")

        result = _korf_excel_status(kdf_path)
        assert result.korf_excel_path is None
        assert result.is_stale is False


# ── Data Class Tests ────────────────────────────────────────────────────────


class TestDataClasses:
    """Tests for parser data classes."""

    def test_case_info_frozen(self):
        """CaseInfo is frozen (immutable)."""
        info = CaseInfo(number="1", name="Rated")
        with pytest.raises(AttributeError):
            info.number = "2"

    def test_pipe_data_defaults(self):
        """PipeData has None defaults for optional fields."""
        pipe = PipeData(name="P-001")
        assert pipe.length is None
        assert pipe.velocity_in is None
        assert pipe.dp_length_criteria_max is None

    def test_pump_data_with_npsha(self):
        """PumpData stores NPSHa value."""
        pump = PumpData(name="P-001", npsha=9.5, shutoff_pressure=18.6)
        assert pump.npsha == 9.5
        assert pump.shutoff_pressure == 18.6

    def test_orifice_data(self):
        """OrificeData stores orifice fields."""
        orif = OrificeData(
            name="FT-001",
            type="Orifice",
            bore=140.5,
            dp_flange_tap=0.189,
            dp_pipe_tap=0.1,
        )
        assert orif.bore == 140.5
        assert orif.dp_flange_tap == 0.189


# ── MultiCaseSummaryBuilder Tests ───────────────────────────────────────────


class TestMultiCaseSummaryBuilder:
    """Tests for MultiCaseSummaryBuilder governing case logic."""

    @pytest.fixture
    def model(self):
        return Model.load(PUMP_KDF)

    @pytest.fixture
    def korf_excel(self, tmp_path):
        return _create_korf_excel(tmp_path, num_cases=2)

    def test_pipe_governing_case_based_on_velocity(self, model, korf_excel):
        """Pipe governing case is determined by max velocity."""
        from pykorf.core.reports.korf_reporter import KorfReporter
        from pykorf.core.reports.multi_case_summary import MultiCaseSummaryBuilder

        reporter = KorfReporter(excel_path=korf_excel, model=model)
        case_data = reporter._get_case_data()
        builder = MultiCaseSummaryBuilder(case_data, model, reporter)

        pipe_df = builder.build_pipe_summary_df()
        assert not pipe_df.empty
        assert "Governing Case" in pipe_df.columns
        assert "Pipe Name" in pipe_df.columns

        for row_idx, row in pipe_df.iterrows():
            governing_case = row.get("Governing Case")
            assert governing_case in ["Rated", "Case2"]

    def test_pump_governing_case_is_first_case(self, model, korf_excel):
        """Pump governing case is always the first case."""
        from pykorf.core.reports.korf_reporter import KorfReporter
        from pykorf.core.reports.multi_case_summary import MultiCaseSummaryBuilder

        reporter = KorfReporter(excel_path=korf_excel, model=model)
        case_data = reporter._get_case_data()
        builder = MultiCaseSummaryBuilder(case_data, model, reporter)

        pump_df = builder.build_pump_summary_df()
        assert not pump_df.empty
        assert "Governing Case" in pump_df.columns

        for row_idx, row in pump_df.iterrows():
            assert row.get("Governing Case") == "Rated"

    def test_valve_per_element_tables(self, model, korf_excel):
        """Valve data is structured as per-element tables with case columns."""
        from pykorf.core.reports.korf_reporter import KorfReporter
        from pykorf.core.reports.multi_case_summary import MultiCaseSummaryBuilder

        reporter = KorfReporter(excel_path=korf_excel, model=model)
        case_data = reporter._get_case_data()
        builder = MultiCaseSummaryBuilder(case_data, model, reporter)

        valve_data = builder.build_valve_per_element_data()
        assert len(valve_data) >= 1

        for valve_info in valve_data:
            assert "valve_name" in valve_info
            assert "parameters" in valve_info
            assert "case_values" in valve_info
            assert valve_info["parameters"] == [
                "Flowrate",
                "Inlet Pressure",
                "Differential Pressure",
            ]
            assert "Rated" in valve_info["case_values"]
            assert "Case2" in valve_info["case_values"]

    def test_placeholder_methods_return_empty_df(self, model, korf_excel):
        """Placeholder methods for other elements return empty DataFrames."""
        from pykorf.core.reports.korf_reporter import KorfReporter
        from pykorf.core.reports.multi_case_summary import MultiCaseSummaryBuilder

        reporter = KorfReporter(excel_path=korf_excel, model=model)
        case_data = reporter._get_case_data()
        builder = MultiCaseSummaryBuilder(case_data, model, reporter)

        assert builder.build_compressor_summary_df().empty
        assert builder.build_exchanger_summary_df().empty
        assert builder.build_misc_summary_df().empty
        assert builder.build_feed_summary_df().empty
        assert builder.build_product_summary_df().empty
        assert builder.build_orifice_summary_df().empty
